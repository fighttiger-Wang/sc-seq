#!/usr/bin/env python3
"""Regression test for the qualitative subcluster workbook contract."""

import argparse
import hashlib
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--work-dir", required=True)
    args = parser.parse_args()
    work = Path(args.work_dir).resolve()
    work.mkdir(parents=True, exist_ok=True)
    repo = next(parent for parent in work.parents if (parent / "local-marketplace").is_dir()) / "local-marketplace"
    skill = repo / "plugins" / "sc-marker-cluster-annotation-auto" / "skills" / "sc-marker-cluster-annotation-auto"
    evidence = {
        "clusters": ["3", "0", "1"], "average_shape": [8, 3], "average_reader": "test",
        "confirmed_metadata": {"species": "Human", "tissue": "fetal lung", "annotation_level": "subcluster", "parent_population": "T_NK"},
        "source_paths": {"cell_avg_exp": "avg_expr_matrix.tsv", "marker_table": "Markergene_list.xlsx", "umap": "umap.png"},
        "cluster_profiles": {
            "0": {"top_markers": [{"gene": "CCR7", "mean_expr": 2.4, "pct1": 0.7, "pct2": 0.2, "log2FC": 1.4}]},
            "1": {"top_markers": [{"gene": "FGFBP2", "mean_expr": 4.67, "pct1": 0.702, "pct2": 0.143, "log2FC": 1.74}]},
            "3": {"top_markers": [{"gene": "TRDC", "mean_expr": 3.1, "pct1": 0.65, "pct2": 0.05, "log2FC": 2.0}]},
        },
        "qualitative_annotation_evidence": {
            "0": {"stable_id": "Tn", "primary_program": "Tn", "candidate_program_audits": [{"label": "Tn", "program_gate": "通过"}], "qualitative_gates": {"identity_anchor": "通过", "sibling_competition": "通过"}},
            "1": {"stable_id": "NK_cell", "primary_program": "NK_cell", "candidate_program_audits": [{"label": "NK_cell", "program_gate": "通过"}], "state_program": [{"program": "activation", "status": "通过", "marker_count": 3}], "qualitative_gates": {"identity_anchor": "通过", "state_program": "通过"}},
            "3": {"stable_id": "gdT", "primary_program": "gdT", "candidate_program_audits": [{"label": "gdT", "program_gate": "通过"}], "qualitative_gates": {"identity_anchor": "通过", "sibling_competition": "通过"}},
        },
        "annotation_evidence_policy": {"decision_model": "qualitative_biological_gates"},
    }
    rows = [("3", "γδT细胞", "gdT", "TRDC"), ("0", "初始T细胞", "Tn", "CCR7"), ("1", "NK细胞", "NK_cell", "FGFBP2")]
    records = [{
        "cluster_id": cluster, "celltype_cn": cn, "celltype_en": en, "stable_id": en,
        "broad_type": "T_NK", "supporting_markers": marker, "candidate_labels": en,
        "rationale": "Sibling-level qualitative program retained.",
        "review_action": "Validate with complementary markers and UMAP topology.",
        "literature_source": "Curated T/NK atlas reference",
    } for cluster, cn, en, marker in rows]
    ep, rp, up, output = work / "evidence.json", work / "records.json", work / "umap_audit.json", work / "subcluster.xlsx"
    ep.write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    rp.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    up.write_text(json.dumps({"clusters": {
        cluster: {
            "reviewed": True,
            "topology_summary": f"Cluster {cluster} reviewed",
            "nearest_clusters": [],
            "marker_umap_relation": "concordant",
            "research_required": False,
            "research_status": "not_required",
            "conflict_resolution_basis": "none",
            "evidence_ids": [],
            "review_action": "retain",
            "identity_action": "retain",
            "provisional_label": en,
            "resolved_label": en,
            "same_label_clusters": [],
            "same_label_topology": "not_applicable",
            "separation_explanation": "none",
            "separation_evidence": "",
        }
        for cluster, _, en, _ in rows
    }}, ensure_ascii=False, indent=2), encoding="utf-8")
    completed = subprocess.run([
        sys.executable, str(skill / "scripts" / "build_annotation_workbook.py"),
        "--records", str(rp), "--evidence", str(ep), "--umap-audit", str(up), "--output", str(output),
        "--workspace-root", str(work.parents[1]), "--force",
    ], text=True, capture_output=True)
    if completed.returncode:
        raise RuntimeError(completed.stdout + completed.stderr)
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["绘图列表", "注释结果", "详细证据", "细胞类型与文献", "说明与数据来源"]
    assert [workbook["绘图列表"].cell(row, 1).value for row in range(2, 5)] == ["0", "1", "3"]
    result_headers = {cell.value: cell.column for cell in workbook["注释结果"][1]}
    assert "质量评分" not in result_headers and "置信度" not in result_headers
    assert workbook["注释结果"].cell(3, result_headers["中文名称"]).fill.fgColor.rgb == "FFF8696B"
    evidence_headers = {cell.value: cell.column for cell in workbook["详细证据"][1]}
    marker_text = workbook["详细证据"].cell(3, evidence_headers["支持 Marker 证据"]).value
    assert "FGFBP2(mean=4.67, ratio=70.20%" in marker_text
    assert all(sheet.auto_filter.ref is None for sheet in workbook.worksheets)

    # A label that disagrees with the qualitative core must not be made
    # acceptable by copying the same label into the workbook-facing record.
    bad_records = json.loads(json.dumps(records))
    bad_records[1]["stable_id"] = "gdT"
    bad_records[1]["celltype_en"] = "gdT"
    bad_records_path = work / "bad_records.json"
    bad_records_path.write_text(json.dumps(bad_records, ensure_ascii=False), encoding="utf-8")
    rejected = subprocess.run([
        sys.executable, str(skill / "scripts" / "build_annotation_workbook.py"),
        "--records", str(bad_records_path), "--evidence", str(ep), "--umap-audit", str(up),
        "--output", str(work / "rejected.xlsx"), "--workspace-root", str(work.parents[1]), "--force",
    ], text=True, capture_output=True)
    assert rejected.returncode != 0

    myeloid_evidence = {
        "clusters": ["B2", "A1"],
        "average_shape": [12, 2],
        "average_reader": "test",
        "confirmed_metadata": {
            "species": "Human", "tissue": "synthetic immune tissue", "annotation_level": "subcluster",
            "parent_population": "Myeloid",
        },
        "source_paths": {
            "cell_avg_exp": "avg_expr_matrix.tsv", "marker_table": "Markergene_list.xlsx", "umap": "umap.png",
        },
        "cluster_profiles": {
            "A1": {"top_markers": [
                {"gene": "CD14", "expr_ratio": 0.88}, {"gene": "FCN1", "expr_ratio": 0.91},
                {"gene": "VCAN", "expr_ratio": 0.86}, {"gene": "S100A8", "expr_ratio": 0.83},
                {"gene": "S100A9", "expr_ratio": 0.82}, {"gene": "CD1C", "expr_ratio": 0.14},
                {"gene": "CLEC10A", "expr_ratio": 0.19},
            ]},
            "B2": {"top_markers": [
                {"gene": "CD1C", "expr_ratio": 0.58}, {"gene": "CLEC10A", "expr_ratio": 0.52},
                {"gene": "FCER1A", "expr_ratio": 0.49}, {"gene": "FCN1", "expr_ratio": 0.35},
            ]},
        },
        "qualitative_annotation_evidence": {
            "A1": {"stable_id": "DC3", "primary_program": "DC3", "candidate_program_audits": [{"label": "DC3", "program_gate": "通过"}, {"label": "Classical_monocyte", "program_gate": "通过"}], "candidate_labels": ["DC3", "Classical_monocyte"]},
            "B2": {"stable_id": "DC3", "primary_program": "DC3", "candidate_program_audits": [{"label": "DC3", "program_gate": "通过"}, {"label": "cDC2", "program_gate": "通过"}], "candidate_labels": ["DC3", "cDC2"]},
        },
        "annotation_evidence_policy": {"decision_model": "qualitative_biological_gates"},
    }
    myeloid_records = [
        {
            "cluster_id": "A1", "celltype_cn": "DC3细胞", "celltype_en": "DC3", "stable_id": "DC3",
            "broad_type": "Myeloid", "supporting_markers": ["CD1C", "CLEC10A", "CD14", "FCN1", "VCAN"],
            "candidate_labels": ["DC3", "Classical_monocyte"],
            "rationale": "Expression-stage provisional DC3 boundary candidate.",
            "review_action": "Integrate the complete UMAP topology.",
            "literature_source": "Curated myeloid references",
        },
        {
            "cluster_id": "B2", "celltype_cn": "DC3细胞", "celltype_en": "DC3", "stable_id": "DC3",
            "broad_type": "Myeloid", "supporting_markers": ["CD1C", "CLEC10A", "FCER1A", "FCN1"],
            "candidate_labels": ["DC3", "cDC2"],
            "rationale": "DC and monocyte-linked programs remain compatible with DC3.",
            "review_action": "Retain after topology review.",
            "literature_source": "Curated myeloid references",
        },
    ]
    myeloid_audit = {"clusters": {
        "A1": {
            "reviewed": True,
            "topology_summary": "The boundary cluster lies on a marker-supported monocyte continuum and is separated from the DC island.",
            "nearest_clusters": ["B2"],
            "marker_umap_relation": "conflict",
            "conflict_reason": "The provisional DC3 label conflicts with the monocyte-continuous topology.",
            "research_required": True,
            "research_status": "reused",
            "conflict_resolution_basis": "integrated_marker_umap_reassessment",
            "evidence_ids": ["synthetic_monocyte_program", "synthetic_umap_topology"],
            "review_action": "Reject provisional DC3 and reassign to the marker-supported monocyte sibling.",
            "identity_action": "reject_and_reassign",
            "provisional_label": "DC3",
            "resolved_label": "Classical_monocyte",
            "resolved_label_cn": "经典单核细胞",
            "reassessment_rationale": "UMAP places the boundary population on the monocyte branch, while a complete monocyte program dominates and the DC signal is limited.",
            "reassessment_marker_support": ["CD14", "FCN1", "VCAN", "S100A8", "S100A9"],
            "same_label_clusters": [],
            "same_label_topology": "not_applicable",
            "separation_explanation": "none",
            "separation_evidence": "",
        },
        "B2": {
            "reviewed": True,
            "topology_summary": "The comparison population remains in a DC-enriched island and is not continuous with the monocyte branch.",
            "nearest_clusters": ["A1"],
            "marker_umap_relation": "concordant",
            "research_required": False,
            "research_status": "not_required",
            "conflict_resolution_basis": "none",
            "evidence_ids": [],
            "review_action": "Retain DC3.",
            "identity_action": "retain",
            "provisional_label": "DC3",
            "resolved_label": "DC3",
            "same_label_clusters": [],
            "same_label_topology": "not_applicable",
            "separation_explanation": "none",
            "separation_evidence": "",
        },
    }}
    myeloid_ep = work / "myeloid_evidence.json"
    myeloid_rp = work / "myeloid_records.json"
    myeloid_up = work / "myeloid_umap_audit.json"
    myeloid_output = work / "myeloid_subcluster.xlsx"
    myeloid_ep.write_text(json.dumps(myeloid_evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    myeloid_rp.write_text(json.dumps(myeloid_records, ensure_ascii=False, indent=2), encoding="utf-8")
    myeloid_up.write_text(json.dumps(myeloid_audit, ensure_ascii=False, indent=2), encoding="utf-8")
    completed = subprocess.run([
        sys.executable, str(skill / "scripts" / "build_annotation_workbook.py"),
        "--records", str(myeloid_rp), "--evidence", str(myeloid_ep), "--umap-audit", str(myeloid_up),
        "--output", str(myeloid_output), "--workspace-root", str(work.parents[1]), "--force",
    ], text=True, capture_output=True)
    if completed.returncode:
        raise RuntimeError(completed.stdout + completed.stderr)
    myeloid_book = load_workbook(myeloid_output)
    plot_rows = {
        str(myeloid_book["绘图列表"].cell(row, 1).value): myeloid_book["绘图列表"].cell(row, 2).value
        for row in range(2, myeloid_book["绘图列表"].max_row + 1)
    }
    assert plot_rows == {"A1": "Classical_monocyte", "B2": "DC3"}
    result_headers = {cell.value: cell.column for cell in myeloid_book["注释结果"][1]}
    result_rows = {
        str(myeloid_book["注释结果"].cell(row, 1).value): row
        for row in range(2, myeloid_book["注释结果"].max_row + 1)
    }
    assert myeloid_book["注释结果"].cell(result_rows["A1"], result_headers["中文名称"]).value == "经典单核细胞"
    assert myeloid_book["注释结果"].cell(result_rows["B2"], result_headers["Celltype_EN"]).value == "DC3"

    interim = work / "myeloid_mapping.tsv"
    completed = subprocess.run([
        sys.executable, str(skill / "scripts" / "build_interim_mapping.py"),
        "--records", str(myeloid_rp), "--evidence", str(myeloid_ep), "--umap-audit", str(myeloid_up),
        "--output", str(interim), "--workspace-root", str(work.parents[1]), "--force",
    ], text=True, capture_output=True)
    if completed.returncode:
        raise RuntimeError(completed.stdout + completed.stderr)
    assert "A1\tClassical_monocyte" in interim.read_text(encoding="utf-8-sig")

    unresolved = json.loads(json.dumps(myeloid_audit))
    unresolved["clusters"]["A1"].pop("identity_action")
    unresolved_path = work / "myeloid_umap_unresolved.json"
    unresolved_path.write_text(json.dumps(unresolved, ensure_ascii=False, indent=2), encoding="utf-8")
    blocked = subprocess.run([
        sys.executable, str(skill / "scripts" / "build_annotation_workbook.py"),
        "--records", str(myeloid_rp), "--evidence", str(myeloid_ep), "--umap-audit", str(unresolved_path),
        "--output", str(work / "blocked.xlsx"), "--workspace-root", str(work.parents[1]), "--force",
    ], text=True, capture_output=True)
    assert blocked.returncode != 0
    assert "requires identity_action" in blocked.stderr

    delivery_dir = work / "delivery"
    delivered = subprocess.run([
        sys.executable, str(skill / "scripts" / "copy_final_workbook.py"),
        "--source", str(myeloid_output), "--workspace-root", str(work.parents[1]),
        "--destination", str(delivery_dir),
    ], text=True, capture_output=True)
    if delivered.returncode:
        raise RuntimeError(delivered.stdout + delivered.stderr)
    assert (delivery_dir / myeloid_output.name).is_file()

    legacy = work / "legacy_four_sheet.xlsx"
    legacy_book = Workbook()
    legacy_book.active.title = "注释结果"
    for name in ("详细证据", "细胞类型与文献", "说明与数据来源"):
        legacy_book.create_sheet(name)
    legacy_book["注释结果"].append(["Cluster", "Confidence", "Quality_score"])
    legacy_book.save(legacy)
    legacy.with_suffix(".qa.json").write_text(json.dumps({
        "status": "pass", "workbook_sha256": hashlib.sha256(legacy.read_bytes()).hexdigest(),
    }), encoding="utf-8")
    legacy_copy = subprocess.run([
        sys.executable, str(skill / "scripts" / "copy_final_workbook.py"),
        "--source", str(legacy), "--workspace-root", str(work.parents[1]),
        "--destination", str(delivery_dir),
    ], text=True, capture_output=True)
    assert legacy_copy.returncode != 0
    assert "violates the five-sheet contract" in legacy_copy.stdout
    print(json.dumps({"status": "pass", "checks": 20, "output": str(output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
