import argparse
import hashlib
import json
import os
import shutil
import stat
import sys
from pathlib import Path

from openpyxl import load_workbook


def load_contract_module():
    local = Path(__file__).resolve().parent
    if (local / "qualitative_annotation_workbook.py").is_file():
        if str(local) not in sys.path:
            sys.path.insert(0, str(local))
        import qualitative_annotation_workbook as module
        return module
    for parent in Path(__file__).resolve().parents:
        shared = parent / "shared" / "sc-annotation-evidence-core"
        if (shared / "qualitative_annotation_workbook.py").is_file():
            if str(shared) not in sys.path:
                sys.path.insert(0, str(shared))
            import qualitative_annotation_workbook as module
            return module
    raise RuntimeError("Shared qualitative workbook contract module not found")


CONTRACT = load_contract_module()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def require_e_drive(path: Path, role: str) -> None:
    if path.drive.upper() != "E:":
        raise ValueError(f"{role} must be an absolute E-drive path: {path}")


def require_inside(path: Path, root: Path, role: str) -> None:
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"{role} must be inside workspace root {root}: {path}") from exc


def is_reparse(path: Path) -> bool:
    if not path.exists():
        return False
    attrs = getattr(os.stat(path, follow_symlinks=False), "st_file_attributes", 0)
    return bool(attrs & stat.FILE_ATTRIBUTE_REPARSE_POINT)


def has_reparse_component(path: Path) -> bool:
    current = path.absolute()
    while True:
        if is_reparse(current):
            return True
        if current.parent == current:
            return False
        current = current.parent


def infer_destination(inputs: list[str]) -> Path:
    if not inputs:
        raise ValueError("destination cannot be inferred because no original input paths were provided")
    resolved_inputs = [Path(value).resolve() for value in inputs]
    for input_path in resolved_inputs:
        require_e_drive(input_path, "original input")
        if not input_path.is_file():
            raise FileNotFoundError(f"original input path is missing: {input_path}")
    parents = {input_path.parent for input_path in resolved_inputs}
    if len(parents) != 1:
        raise ValueError(f"destination cannot be inferred from multiple input directories: {sorted(map(str, parents))}")
    return parents.pop()


def validate_formal_workbook(source: Path) -> dict:
    qa_path = source.with_suffix(".qa.json")
    if not qa_path.is_file():
        raise ValueError(f"formal workbook lacks builder QA sidecar: {qa_path}")
    qa = json.loads(qa_path.read_text(encoding="utf-8"))
    if qa.get("status") != "pass":
        raise ValueError(f"formal workbook QA status is not pass: {qa.get('status')}")
    actual_hash = sha256(source)
    if qa.get("workbook_sha256") != actual_hash:
        raise ValueError("formal workbook hash does not match its builder QA sidecar")

    workbook = load_workbook(source, read_only=False, data_only=False)
    expected = ["绘图列表", "注释结果", "详细证据", "细胞类型与文献", "说明与数据来源"]
    if workbook.sheetnames != expected:
        raise ValueError(f"formal workbook violates the five-sheet contract: {workbook.sheetnames}")
    headers = {
        "绘图列表": CONTRACT.PLOT_HEADERS,
        "注释结果": CONTRACT.RESULT_HEADERS,
        "详细证据": CONTRACT.EVIDENCE_HEADERS,
        "细胞类型与文献": CONTRACT.LITERATURE_HEADERS,
        "说明与数据来源": CONTRACT.SOURCE_HEADERS,
    }
    freezes = {"绘图列表": "A2", "注释结果": "D2", "详细证据": "D2", "细胞类型与文献": "B2", "说明与数据来源": "A2"}
    forbidden = {"Confidence", "Quality_score", "置信度", "质量评分", "候选评分", "评分差值"}
    red_rgb = {"FFF8696B", "F8696B"}
    for name in expected:
        sheet = workbook[name]
        observed_headers = [cell.value for cell in sheet[1]]
        if observed_headers != headers[name]:
            raise ValueError(f"{name} header contract mismatch")
        if forbidden.intersection(str(value) for value in observed_headers if value is not None):
            raise ValueError(f"{name} contains forbidden score/confidence headers")
        if sheet.auto_filter.ref is not None:
            raise ValueError(f"{name} must not enable autofilter")
        if str(sheet.freeze_panes) != freezes[name]:
            raise ValueError(f"{name} freeze panes mismatch: {sheet.freeze_panes}")
        for row in sheet.iter_rows():
            if row[0].row > 1 and sheet.row_dimensions[row[0].row].height is None:
                raise ValueError(f"{name} row {row[0].row} lacks fixed height")
            for cell in row:
                if cell.alignment.wrap_text or cell.alignment.shrink_to_fit:
                    raise ValueError(f"{name}!{cell.coordinate} enables wrapping or shrink-to-fit")
                rgb = str(cell.fill.fgColor.rgb or "").upper()
                if rgb in red_rgb and not (name == "注释结果" and cell.column == 2 and cell.row > 1):
                    raise ValueError(f"red warning fill is outside 注释结果/中文名称: {name}!{cell.coordinate}")
    return qa


def main() -> None:
    parser = argparse.ArgumentParser(description="Copy only a validated final annotation workbook to an E-drive destination.")
    parser.add_argument("--source", required=True)
    parser.add_argument("--workspace-root", required=True)
    parser.add_argument("--destination")
    parser.add_argument("--input", action="append", default=[], help="Original input file; repeat for paired inputs to infer their common directory.")
    args = parser.parse_args()
    workspace_root = Path(args.workspace_root).resolve()
    source = Path(args.source).resolve()
    destination_arg = Path(args.destination) if args.destination else infer_destination(args.input)
    require_e_drive(workspace_root, "workspace root")
    require_e_drive(source, "source")
    require_inside(source, workspace_root, "source")
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise ValueError(f"source must be an existing .xlsx workbook: {source}")
    qa = validate_formal_workbook(source)
    if not destination_arg.is_absolute():
        raise ValueError(f"destination must be an absolute E-drive path: {destination_arg}")
    require_e_drive(destination_arg, "destination")
    check_dir = destination_arg.parent if destination_arg.suffix.lower() == ".xlsx" else destination_arg
    if has_reparse_component(check_dir):
        raise ValueError(f"destination directory must not contain a junction/reparse point: {check_dir}")
    destination = destination_arg if destination_arg.suffix.lower() == ".xlsx" else destination_arg / source.name
    destination = destination.absolute()
    destination_dir = destination.parent
    if destination.suffix.lower() != ".xlsx":
        raise ValueError(f"destination workbook must end in .xlsx: {destination}")
    destination_dir.mkdir(parents=True, exist_ok=True)
    if has_reparse_component(destination_dir):
        raise ValueError(f"destination directory must not contain a junction/reparse point: {destination_dir}")
    source_hash = sha256(source)
    if source.resolve() != destination.resolve():
        shutil.copy2(source, destination)
    destination_hash = sha256(destination)
    if source_hash != destination_hash:
        raise RuntimeError("destination hash does not match source after copy")
    print(json.dumps({
        "status": "copied", "source": str(source), "destination": str(destination),
        "destination_mode": "explicit" if args.destination else "inferred_from_original_inputs",
        "sha256": source_hash, "qa_status": qa.get("status"), "copied_files": 1,
    }, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except (FileNotFoundError, ValueError) as exc:
        print(json.dumps({"status": "needs_delivery_path", "error": str(exc)}, ensure_ascii=False))
        raise SystemExit(2)
