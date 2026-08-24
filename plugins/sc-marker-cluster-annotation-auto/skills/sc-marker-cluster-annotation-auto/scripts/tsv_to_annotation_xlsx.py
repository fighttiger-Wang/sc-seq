#!/usr/bin/env python3
import argparse
import csv
import os
import platform
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def require_e_output(path):
    resolved = Path(path).resolve()
    if platform.system() == "Windows" and os.path.splitdrive(str(resolved))[0].upper() != "E:":
        raise ValueError(f"Output must be on E: {resolved}")
    return resolved


def read_tsv(path):
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle, delimiter="\t"))


def write_xlsx(rows, output, sheet_name):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in range(1, sheet.max_column + 1):
        sample = [str(sheet.cell(row, column).value or "") for row in range(1, min(sheet.max_row, 200) + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max(map(len, sample), default=8) + 2, 10), 45)
    workbook.save(output)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--avg-tsv", required=True)
    parser.add_argument("--markers-tsv", required=True)
    parser.add_argument("--output-dir", required=True)
    args = parser.parse_args()
    output_dir = require_e_output(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    avg_output, marker_output = output_dir / "cell_avg_exp.xlsx", output_dir / "Markergene_list.xlsx"
    write_xlsx(read_tsv(args.avg_tsv), avg_output, "Sheet1")
    write_xlsx(read_tsv(args.markers_tsv), marker_output, "Sheet1")
    print(f"Created:\n{avg_output}\n{marker_output}")


if __name__ == "__main__":
    main()
