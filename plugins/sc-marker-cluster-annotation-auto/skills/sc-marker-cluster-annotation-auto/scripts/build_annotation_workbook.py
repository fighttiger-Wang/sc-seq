#!/usr/bin/env python3
"""Build and structurally verify the standardized annotation workbook."""

import argparse
import json
import os
import platform
import re
import subprocess
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from umap_audit import apply_umap_audit, load_umap_audit, validate_umap_audit


SKILL_NAME = "sc-marker-cluster-annotation-auto"
SKILL_VERSION = "0.6.0"


def nonproduction_output(output):
    for directory in (output.parent, *output.parents):
        name = directory.name
        if (
            re.search(r"(^|[_-])(test|regression|validation|builder|preflight)([_-]|$)", name, re.I)
            or name.lower().startswith("skill_")
        ):
            return True
    return False


def register_shared_case(workspace, records_path, evidence_path, qa_path, output):
    if nonproduction_output(output):
        return {"status": "skipped_nonproduction", "project_dir": str(output.parent)}
    script = workspace / "local-marketplace" / "shared" / "sc-annotation-case-registry" / "case_registry.py"
    sidecar = output.with_suffix(".case-registry.json")
    if not script.is_file():
        result = {"status": "failed", "error": f"Shared case registry script missing: {script}"}
    else:
        completed = subprocess.run(
            [sys.executable, str(script), "register", "--records", str(records_path), "--evidence", str(evidence_path),
             "--qa", str(qa_path), "--project-dir", str(output.parent), "--skill-name", SKILL_NAME,
             "--skill-version", SKILL_VERSION], text=True, capture_output=True
        )
        if completed.returncode == 0:
            try:
                result = json.loads(completed.stdout)
            except json.JSONDecodeError:
                result = {"status": "failed", "error": "Registry returned invalid JSON", "stdout": completed.stdout[-2000:]}
        else:
            result = {"status": "failed", "error": completed.stderr.strip() or completed.stdout.strip(), "returncode": completed.returncode}
    sidecar.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return result


def deliver_final_workbook(workspace, output, evidence):
    if nonproduction_output(output):
        return {"status": "skipped_nonproduction"}
    paths = evidence.get("source_paths", {})
    original_inputs = [
        str(paths.get(key, "")).strip()
        for key in ("cell_avg_exp", "marker_table")
        if str(paths.get(key, "")).strip()
    ]
    if len(original_inputs) != 2:
        raise RuntimeError(
            "Formal production delivery requires the original average-expression and marker-table paths"
        )
    script = Path(__file__).with_name("copy_final_workbook.py")
    command = [
        sys.executable, str(script), "--source", str(output),
        "--workspace-root", str(workspace),
    ]
    for input_path in original_inputs:
        command.extend(["--input", input_path])
    completed = subprocess.run(command, text=True, capture_output=True)
    if completed.returncode != 0:
        raise RuntimeError(
            "Final workbook passed QA but automatic delivery to the original input directory failed: "
            + (completed.stderr.strip() or completed.stdout.strip())
        )
    try:
        result = json.loads(completed.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError("Final delivery returned invalid JSON") from exc
    output.with_suffix(".delivery.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return result


FIELDS = [
    "cluster_id", "celltype_cn", "celltype_en", "broad_type", "fine_type", "state",
    "supporting_markers", "conflicting_markers", "candidate_labels", "confidence",
    "quality_score", "mixed_or_doublet", "mixture_type", "possible_components",
    "rationale", "manual_review", "review_action",
    "evidence_mode", "evidence_completeness", "primary_evidence_label", "primary_evidence_score",
    "runner_up_evidence_label", "runner_up_evidence_score", "score_margin", "positive_marker_coverage",
    "detection_specificity", "rival_lineage", "rival_lineage_score", "risk_level",
    "recommended_action", "decision_trace",
]
HEADERS = [
    "Cluster", "建议中文名称", "English label", "大类谱系", "精细类型", "细胞状态",
    "关键 Marker", "冲突 Marker", "候选标签", "置信度", "质量评分", "疑似多细胞/双细胞",
    "混合类型", "可能组成", "判定依据", "人工复核", "人工复核建议",
    "证据模式", "证据完整度", "确定性主候选", "主候选评分", "确定性次候选", "次候选评分",
    "评分差", "核心Marker覆盖", "检测特异性", "竞争谱系", "竞争谱系评分", "风险等级",
    "处置建议", "确定性判定轨迹",
]
STRUCTURED_FIELDS = [
    "stable_id", "formal_identity_fallback", "parent_path", "tissue_module",
    "expected_parent_id", "off_parent_detected", "off_parent_reassignment", "off_parent_candidate", "off_parent_candidate_score",
    "developmental_stage", "ontology_node_kind", "tissue_scope", "tissue_scope_match", "tissue_context_review",
    "disease_role", "state_list", "primary_state",
    "cross_species_inference", "panel_species", "marker_panel_evidence_ids", "display_label",
    "mixed_population", "suspected_doublet", "auto_merge_allowed",
]
STRUCTURED_HEADERS = [
    "Stable_ID", "Formal_Identity_Fallback", "Parent_Path", "Tissue_Module",
    "Expected_Parent_ID", "Off_Parent_Detected", "Off_Parent_Reassignment", "Off_Parent_Candidate", "Off_Parent_Candidate_Score",
    "Developmental_Stage", "Ontology_Node_Kind", "Tissue_Scope", "Tissue_Scope_Match", "Tissue_Context_Review",
    "Disease_Role", "State_List", "Primary_State",
    "Cross_Species_Inference", "Panel_Species", "Marker_Panel_Evidence", "Display_Label",
    "Mixed_Population", "Suspected_Doublet", "Auto_Merge_Allowed",
]
FIELDS[6:6] = STRUCTURED_FIELDS
HEADERS[6:6] = STRUCTURED_HEADERS
AUDIT_FIELDS = [
    "label_basis", "canonical_subtype", "top_marker_gene", "literature_source",
    "naming_grammar", "contextually_excluded_naming_markers",
]
LABEL_BASIS = {
    "canonical_subtype", "validated_external_candidate", "researched_branch_fallback",
    "top_marker_fallback", "feature_gene_fallback",
}
CONFIDENCE = {"high", "medium-high", "medium", "low"}
FINAL_LABEL_INVALID = re.compile(r"[^A-Za-z0-9_\u3400-\u4DBF\u4E00-\u9FFF]+")
REDUNDANT_T_SUFFIX = re.compile(
    r"(^|_)(Tn|Tcm|Tem|Trm|Tex|Treg|Tfh|Th1|Th2|Th17|gdT|NKT|MAIT)_T$",
    re.IGNORECASE,
)

MAIN_FIELDS = [
    "cluster_id", "celltype_cn", "celltype_en", "broad_type", "developmental_stage", "state",
    "disease_role", "supporting_markers", "conflicting_markers", "candidate_labels", "confidence",
    "quality_score", "mixed_or_doublet", "mixture_type", "possible_components", "rationale",
    "manual_review", "review_action",
]
MAIN_HEADERS = [
    "Cluster", "中文名称", "Celltype_EN", "大类谱系", "发育阶段", "细胞状态",
    "疾病或组织角色", "关键 Marker", "冲突 Marker", "候选标签", "置信度",
    "质量评分", "疑似多细胞/双细胞", "混合类型", "可能组成", "判定依据",
    "需要复核", "复核建议",
]
EVIDENCE_FIELDS = [
    "cluster_id", "stable_id", "fine_type", "formal_identity_fallback", "parent_path", "tissue_module",
    "expected_parent_id", "off_parent_detected", "off_parent_reassignment", "off_parent_candidate",
    "off_parent_candidate_score", "ontology_node_kind", "tissue_scope", "tissue_scope_match",
    "tissue_context_review", "state_list", "primary_state", "cross_species_inference", "panel_species",
    "marker_panel_evidence_ids", "evidence_mode", "evidence_completeness", "primary_evidence_label",
    "primary_evidence_score", "runner_up_evidence_label", "runner_up_evidence_score", "score_margin",
    "positive_marker_coverage", "detection_specificity", "rival_lineage", "rival_lineage_score",
    "risk_level", "recommended_action", "auto_merge_allowed", "label_basis", "canonical_subtype",
    "literature_source", "naming_grammar",
]
EVIDENCE_HEADERS = [
    "Cluster", "Stable_ID", "精细类型", "身份回退原因", "本体路径", "组织模块",
    "预期父群", "检出父群外程序", "父群外重分配", "父群外候选", "父群外候选评分",
    "本体节点类型", "适用组织", "组织匹配", "组织背景复核", "完整状态", "主要状态",
    "跨物种推断", "参考物种", "Marker 证据", "证据模式", "证据完整度", "主候选",
    "主候选评分", "次候选", "次候选评分", "评分差", "核心 Marker 覆盖",
    "检测特异性", "竞争谱系", "竞争谱系评分", "风险等级", "处置建议",
    "允许自动合并", "命名依据", "标准亚型", "文献依据", "命名规则",
]
EVIDENCE_FIELDS.extend([
    "umap_review_status", "umap_neighbors", "umap_topology", "umap_marker_concordance",
    "umap_review_action", "umap_research_status", "umap_evidence_ids",
    "umap_same_label_clusters", "umap_same_label_topology",
    "umap_separation_explanation", "umap_separation_evidence", "umap_conflict_resolution_basis",
    "identity_boundary_audit", "boundary_validation_required", "boundary_validation_resolved",
])
EVIDENCE_HEADERS.extend([
    "UMAP复核状态", "UMAP近邻Cluster", "UMAP拓扑摘要", "Marker与UMAP关系",
    "UMAP复核动作", "UMAP触发研究状态", "UMAP证据ID",
    "同标签Cluster", "同标签拓扑", "空间分离解释", "空间分离证据",
    "Conflict resolution basis", "Identity boundary audit", "Boundary validation required", "Boundary validation resolved",
])


def inject_deterministic_evidence(records, evidence):
    decisions = evidence.get("deterministic_annotation_evidence", {})
    source_registry = evidence.get("annotation_evidence_policy", {}).get("evidence_source_registry", {})
    for record in records:
        cluster = str(record.get("cluster_id", ""))
        if cluster not in decisions:
            raise ValueError(f"Cluster {cluster} lacks deterministic annotation evidence")
        decision = decisions[cluster]
        identity_override = record.get("label_basis") in {
            "validated_external_candidate", "researched_branch_fallback"
        }
        for field in (
            "evidence_mode", "evidence_completeness", "primary_evidence_label", "primary_evidence_score",
            "runner_up_evidence_label", "runner_up_evidence_score", "score_margin", "positive_marker_coverage",
            "detection_specificity", "rival_lineage", "rival_lineage_score", "risk_level", "recommended_action",
        ):
            record[field] = decision[field]
        record["identity_boundary_audit"] = json.dumps(
            decision.get("identity_boundary_audit", {}), ensure_ascii=False, sort_keys=True
        )
        record["boundary_validation_required"] = bool(decision.get("boundary_validation_required", False))
        record["boundary_validation_resolved"] = bool(decision.get("boundary_validation_resolved", False))
        for field in (
            "stable_id", "formal_identity_fallback", "developmental_stage", "ontology_node_kind",
            "expected_parent_id", "off_parent_detected", "off_parent_reassignment", "off_parent_candidate", "off_parent_candidate_score",
            "tissue_scope_match", "tissue_context_review", "primary_state", "cross_species_inference",
            "panel_species", "display_label",
        ):
            if not identity_override or not record.get(field):
                record[field] = decision.get(field, "")
        for field in ("mixed_population", "suspected_doublet", "auto_merge_allowed"):
            record[field] = decision.get(field, False)
        for field in ("parent_path", "tissue_module", "tissue_scope", "disease_role", "state_list", "marker_panel_evidence_ids"):
            if not identity_override or not record.get(field):
                record[field] = json.dumps(decision.get(field, []), ensure_ascii=False)
        if record.get("label_basis") == "canonical_subtype" and not str(record.get("literature_source", "")).strip():
            source_labels = []
            for evidence_id in decision.get("marker_panel_evidence_ids", []):
                source = source_registry.get(evidence_id, {})
                title = str(source.get("title") or source.get("short_name") or evidence_id).strip()
                identifier = str(source.get("doi") or source.get("url_or_path") or evidence_id).strip()
                source_labels.append(f"{title} ({identifier})" if identifier != title else title)
            record["literature_source"] = "; ".join(source_labels)
        if identity_override:
            record["display_label"] = record.get("stable_id", record.get("celltype_en", ""))
            record["manual_review"] = True
            if record.get("stable_id") == "DC3" and decision.get("boundary_validation_required"):
                record["auto_merge_allowed"] = False
                record["mixture_type"] = "unresolved_DC3_monocyte_boundary"
                record["possible_components"] = "DC3; Monocyte"
        if not record.get("state"):
            record["state"] = decision.get("primary_state", "")
        if decision.get("mixed_population"):
            record["mixed_or_doublet"] = True
            record["manual_review"] = True
            record["mixture_type"] = decision.get("mixture_type") or record.get("mixture_type") or "mixed_population/suspected_doublet"
            components = decision.get("possible_components", [])
            if components:
                record["possible_components"] = "; ".join(str(item) for item in components)
        elif decision.get("off_parent_detected"):
            record["manual_review"] = True
            record["mixture_type"] = decision.get("mixture_type") or "off_parent_contaminant"
            components = decision.get("possible_components", [])
            if components:
                record["possible_components"] = "; ".join(str(item) for item in components)
        record["decision_trace"] = json.dumps(decision.get("decision_trace", {}), ensure_ascii=False, sort_keys=True)


def cluster_sort_key(value):
    """Sort numeric cluster IDs by value, then nonnumeric IDs naturally."""
    text = str(value).strip()
    try:
        number = Decimal(text)
        if number.is_finite():
            return (0, number, text)
    except InvalidOperation:
        pass
    natural = tuple((0, int(part)) if part.isdigit() else (1, part.lower()) for part in re.split(r"(\d+)", text))
    return (1, natural, text)


def resolved_e(path, role):
    resolved = Path(path).resolve()
    if platform.system() == "Windows" and os.path.splitdrive(str(resolved))[0].upper() != "E:":
        raise ValueError(f"{role} must be on E: {resolved}")
    return resolved


def normalize_final_label(value):
    """Make labels portable and remove a trailing T already encoded by a subtype token."""
    text = FINAL_LABEL_INVALID.sub("_", str(value).strip())
    text = re.sub(r"_+", "_", text).strip("_")
    text = REDUNDANT_T_SUFFIX.sub(r"\1\2", text)
    if not text:
        raise ValueError(f"Final label becomes empty after normalization: {value!r}")
    return text


def normalize_record_labels(records):
    changes = []
    for record in records:
        for field in ("celltype_cn", "celltype_en"):
            original = str(record[field])
            normalized = normalize_final_label(original)
            record[field] = normalized
            if normalized != original:
                changes.append({
                    "cluster_id": str(record["cluster_id"]),
                    "field": field,
                    "original": original,
                    "normalized": normalized,
                })
    return changes

def within(path, root, role):
    path, root = Path(path).resolve(), Path(root).resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"{role} must be inside workspace root {root}: {path}") from exc
    return path


def structured_list(value):
    if isinstance(value, list):
        return [str(item) for item in value]
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return []
        return [str(item) for item in parsed] if isinstance(parsed, list) else []
    return []


def hierarchy_depth_conflicts(records):
    def is_blocked_mixed_parent_fallback(record):
        return (
            str(record.get("formal_identity_fallback", "")) in {
                "mixed_incompatible_sublineages", "mixed_parent_off_parent_lineages"
            }
            and bool(record.get("mixed_population"))
            and not bool(record.get("auto_merge_allowed", True))
        )

    labels = {
        str(record.get("stable_id", ""))
        for record in records
        if str(record.get("stable_id", ""))
        and not is_blocked_mixed_parent_fallback(record)
    }
    conflicts = []
    for record in records:
        if is_blocked_mixed_parent_fallback(record):
            continue
        child = str(record.get("stable_id", ""))
        ancestors = structured_list(record.get("parent_path", []))[:-1]
        for ancestor in ancestors:
            if ancestor in labels:
                conflicts.append({"ancestor": ancestor, "descendant": child})
    return sorted(conflicts, key=lambda item: (item["ancestor"], item["descendant"]))


def validate(records, clusters, evidence):
    errors = []
    record_clusters = [str(r.get("cluster_id", "")) for r in records]
    expected_clusters = sorted((str(c) for c in clusters), key=cluster_sort_key)
    if record_clusters != expected_clusters:
        errors.append(f"Cluster order/membership mismatch: expected={expected_clusters}, records={record_clusters}")
    top_by_cluster = {}
    informative_by_cluster = {}
    for cluster in clusters:
        profile = evidence["cluster_profiles"][str(cluster)]
        naming_top = profile.get("naming_top_marker")
        top_by_cluster[str(cluster)] = str(naming_top["gene"]) if naming_top else None
        informative_by_cluster[str(cluster)] = [
            str(marker["gene"]) for marker in profile.get("top_informative_markers", [])
        ]
    grammars = set()
    canonical_counts = {}
    for index, record in enumerate(records):
        missing = [field for field in FIELDS + AUDIT_FIELDS if field not in record]
        if missing:
            errors.append(f"Record {index} missing fields: {missing}")
            continue
        cluster_id = str(record["cluster_id"])
        stable_id = str(record.get("stable_id", "")).strip()
        display_label = str(record.get("display_label", "")).strip()
        english_label = str(record.get("celltype_en", "")).strip()
        if stable_id in {"CD4_Tex", "CD8_Tex"}:
            errors.append(f"Cluster {cluster_id} uses deprecated identity/state boundary Stable_ID {stable_id}")
        for label_name, label_value in (("display_label", display_label), ("English label", english_label)):
            if re.fullmatch(r"Exhausted_.+_Tex", label_value):
                errors.append(f"Cluster {cluster_id} {label_name} repeats exhaustion semantics: {label_value}")
        expected_top = top_by_cluster[cluster_id]
        selected_marker = str(record["top_marker_gene"])
        contextual_exclusions = record["contextually_excluded_naming_markers"]
        if record["label_basis"] not in LABEL_BASIS:
            errors.append(f"Cluster {cluster_id} invalid label_basis: {record['label_basis']}")
        if not isinstance(contextual_exclusions, list):
            errors.append(f"Cluster {cluster_id} contextually_excluded_naming_markers must be a list")
            contextual_exclusions = []
        if expected_top is None:
            if record["label_basis"] in {"top_marker_fallback", "feature_gene_fallback"}:
                errors.append(
                    f"Cluster {cluster_id} has no standardized informative marker; "
                    "marker fallback is not allowed and requires a canonical label or manual review"
                )
        elif record["label_basis"] != "feature_gene_fallback" and selected_marker != expected_top:
            errors.append(
                f"Cluster {cluster_id} top_marker_gene must equal highest-log2FC standardized "
                f"informative naming marker {expected_top}"
            )
        if record["label_basis"] == "feature_gene_fallback" and expected_top is not None:
            informative = informative_by_cluster[cluster_id]
            if selected_marker not in informative:
                errors.append(
                    f"Cluster {cluster_id} feature fallback marker {selected_marker} must occur in top_informative_markers"
                )
            else:
                preceding = informative[:informative.index(selected_marker)]
                excluded_genes = []
                for item in contextual_exclusions:
                    if not isinstance(item, dict) or not str(item.get("gene", "")).strip() or not str(item.get("reason", "")).strip():
                        errors.append(f"Cluster {cluster_id} contextual exclusions require non-empty gene and reason")
                        continue
                    excluded_genes.append(str(item["gene"]))
                if excluded_genes != preceding:
                    errors.append(
                        f"Cluster {cluster_id} contextual exclusions must exactly match preceding informative markers: "
                        f"expected={preceding}, recorded={excluded_genes}"
                    )
            if not record["manual_review"]:
                errors.append(f"Cluster {cluster_id} feature_gene_fallback requires manual_review=true")
            if record["confidence"] == "high":
                errors.append(f"Cluster {cluster_id} feature_gene_fallback confidence cannot be high")
        elif contextual_exclusions:
            errors.append(f"Cluster {cluster_id} contextual exclusions are allowed only for feature_gene_fallback")
        if not str(record["naming_grammar"]).strip():
            errors.append(f"Cluster {cluster_id} lacks naming_grammar")
        else:
            grammars.add(str(record["naming_grammar"]).strip())
        canonical = str(record["canonical_subtype"]).strip()
        literature_sources = [
            item.strip() for item in re.split(r"[;；]", str(record["literature_source"])) if item.strip()
        ]
        branch_fallback = str(record.get("formal_identity_fallback", "")) == "branch_identity_no_supported_leaf"
        deterministic_decision = evidence.get("deterministic_annotation_evidence", {}).get(cluster_id, {})
        if deterministic_decision.get("resolution_search_required") and record["label_basis"] not in {
            "validated_external_candidate", "researched_branch_fallback"
        }:
            errors.append(
                f"Cluster {cluster_id} still requires mandatory leaf resolution; "
                "a temporary confirmed-parent mapping cannot be published as the final workbook"
            )
        if str(record.get("formal_identity_fallback", "")) == "confirmed_parent":
            errors.append(
                f"Cluster {cluster_id} retains formal_identity_fallback=confirmed_parent; "
                "resolve it to a leaf/external identity before formal delivery"
            )
        if branch_fallback and record["label_basis"] not in {
            "validated_external_candidate", "researched_branch_fallback"
        }:
            errors.append(
                f"Cluster {cluster_id} branch_identity_no_supported_leaf requires a targeted resolution search; "
                "use a validated external leaf or a two-source researched_branch_fallback before delivery"
            )
        if record["label_basis"] == "canonical_subtype":
            if not canonical:
                errors.append(f"Cluster {cluster_id} canonical_subtype label lacks canonical_subtype")
            if not str(record["literature_source"]).strip():
                errors.append(f"Cluster {cluster_id} canonical subtype lacks literature_source")
            canonical_counts[canonical] = canonical_counts.get(canonical, 0) + 1
            if canonical != str(record.get("stable_id", "")).strip():
                errors.append(
                    f"Cluster {cluster_id} canonical_subtype {canonical} must equal deterministic stable_id {record.get('stable_id')}"
                )
            if str(record.get("celltype_en", "")).strip() != str(record.get("display_label", "")).strip():
                errors.append(
                    f"Cluster {cluster_id} canonical English label must equal deterministic display_label {record.get('display_label')}"
                )
        elif record["label_basis"] == "validated_external_candidate":
            if not canonical or canonical != stable_id:
                errors.append(f"Cluster {cluster_id} validated external candidate requires canonical_subtype=stable_id")
            if len(literature_sources) < 2:
                errors.append(f"Cluster {cluster_id} validated external candidate requires at least two independent sources")
            if not record["manual_review"]:
                errors.append(f"Cluster {cluster_id} validated external candidate requires manual_review=true")
            if record["confidence"] == "high":
                errors.append(f"Cluster {cluster_id} validated external candidate cannot receive high confidence before regression promotion")
            if english_label != stable_id:
                errors.append(f"Cluster {cluster_id} external candidate English label must equal stable_id")
        elif record["label_basis"] == "researched_branch_fallback":
            if not branch_fallback:
                errors.append(
                    f"Cluster {cluster_id} researched_branch_fallback requires "
                    "formal_identity_fallback=branch_identity_no_supported_leaf"
                )
            if not canonical or canonical != stable_id:
                errors.append(f"Cluster {cluster_id} researched branch fallback requires canonical_subtype=stable_id")
            if len(literature_sources) < 2:
                errors.append(f"Cluster {cluster_id} researched branch fallback requires at least two independent sources")
            if not record["manual_review"]:
                errors.append(f"Cluster {cluster_id} researched branch fallback requires manual_review=true")
            if record["confidence"] == "high":
                errors.append(f"Cluster {cluster_id} researched branch fallback confidence cannot be high")
            if not str(record.get("candidate_labels", "")).strip() or not str(record.get("conflicting_markers", "")).strip():
                errors.append(
                    f"Cluster {cluster_id} researched branch fallback requires competing candidates and conflicts"
                )
        elif record["label_basis"] in {"top_marker_fallback", "feature_gene_fallback"}:
            if canonical:
                errors.append(f"Cluster {cluster_id} marker fallback must leave canonical_subtype blank")
            if str(record["literature_source"]).strip():
                errors.append(f"Cluster {cluster_id} marker fallback must leave literature_source blank")
            if expected_top is not None and not str(record["celltype_en"]).startswith(selected_marker + "_"):
                errors.append(f"Cluster {cluster_id} fallback label must start with {selected_marker}_")
        if record["confidence"] not in CONFIDENCE:
            errors.append(f"Cluster {record['cluster_id']} invalid confidence: {record['confidence']}")
        score = record["quality_score"]
        if not isinstance(score, (int, float)) or not 0 <= score <= 100:
            errors.append(f"Cluster {record['cluster_id']} quality_score must be 0-100")
        if not record["celltype_cn"] or not record["celltype_en"] or not record["rationale"]:
            errors.append(f"Cluster {record['cluster_id']} lacks a final label or rationale")
        if record["risk_level"] != "R0_ACCEPT" and not record["manual_review"]:
            errors.append(f"Cluster {cluster_id} deterministic risk {record['risk_level']} requires manual_review=true")
        if str(record.get("ontology_node_kind", "")).endswith("_group"):
            errors.append(f"Cluster {cluster_id} cannot use structural ontology group {record.get('stable_id')} as a final subtype")
        if record.get("tissue_context_review"):
            if not record["manual_review"]:
                errors.append(f"Cluster {cluster_id} noncanonical tissue identity requires manual_review=true")
            if record["confidence"] == "high":
                errors.append(f"Cluster {cluster_id} noncanonical tissue identity confidence cannot be high")
        if record.get("mixed_population") and (record.get("auto_merge_allowed") or not record.get("mixed_or_doublet")):
            errors.append(f"Cluster {cluster_id} mixed population must block automatic merging and carry mixed_or_doublet=true")
        if str(record.get("risk_level", "")).startswith(("R1_", "R2_", "R3_")) and not record.get("manual_review"):
            errors.append(f"Cluster {cluster_id} non-R0 evidence requires manual_review=true")
        boundary = deterministic_decision.get("identity_boundary_audit", {})
        neutrophil_boundary = boundary.get("neutrophil_vs_monocyte", {})
        if stable_id in {"Neutrophil", "Immature_neutrophil", "Neutrophil_progenitor"}:
            if boundary.get("assessed") and not neutrophil_boundary.get("neutrophil_program_passed"):
                errors.append(
                    f"Cluster {cluster_id} {stable_id} label fails the neutrophil program gate; "
                    "CSF3R/FCGR3B alone cannot override a coherent monocyte program"
                )
        dc3_boundary = boundary.get("dc3_vs_monocyte", {})
        if stable_id == "DC3":
            if not dc3_boundary.get("dc3_boundary_candidate"):
                errors.append(
                    f"Cluster {cluster_id} DC3 label lacks the required APC, DC-specific, and monocyte programs"
                )
            if not dc3_boundary.get("cell_level_validated") and record.get("auto_merge_allowed"):
                errors.append(f"Cluster {cluster_id} provisional DC3/monocyte boundary must block automatic merging")
        if deterministic_decision.get("boundary_validation_required") and not deterministic_decision.get("boundary_validation_resolved"):
            errors.append(
                f"Cluster {cluster_id} has an unresolved DC3/monocyte identity boundary; literature review alone "
                "cannot satisfy the required cell-level coexpression or reclustering gate for formal delivery"
            )
        if record.get("umap_same_label_topology") == "disconnected" and record.get("auto_merge_allowed"):
            errors.append(f"Cluster {cluster_id} disconnected repeated identity must block automatic merging")
        if record["evidence_mode"] == "minimal" and record["confidence"] == "high":
            errors.append(f"Cluster {cluster_id} positive-marker-only evidence cannot receive high confidence")
        if "JCHAIN" in str(record["supporting_markers"]).upper() and "plasma" in str(record["celltype_en"]).lower():
            coherent = sum(g in str(record["supporting_markers"]).upper() for g in ["PRDM1", "XBP1", "SDC1", "MZB1", "DERL3"])
            if coherent < 3:
                errors.append(f"Cluster {record['cluster_id']} plasma label lacks coherent secretion program")
    if len(grammars) > 1:
        errors.append(f"Multiple naming grammars in one table: {sorted(grammars)}")
    depth_conflicts = hierarchy_depth_conflicts(records)
    if depth_conflicts:
        errors.append(
            "Final subcluster table mixes ontology ancestors with their descendants; run the mandatory "
            "resolution-search pass and refine the ancestor before delivery: "
            + json.dumps(depth_conflicts, ensure_ascii=False)
        )
    # Repeated canonical labels are valid; cluster_id provides uniqueness.
    if errors:
        raise ValueError("\n".join(errors))


def set_widths(ws, widths):
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def human_value(value):
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (list, tuple, set)):
        return "、".join(human_value(item) for item in value if human_value(item))
    if isinstance(value, dict):
        return "；".join(f"{key}：{human_value(item)}" for key, item in value.items() if human_value(item))
    text = str(value).strip()
    if text.startswith(("[", "{")):
        try:
            return human_value(json.loads(text))
        except json.JSONDecodeError:
            pass
    text = text.replace("[", "").replace("]", "").replace("{", "").replace("}", "")
    return re.sub(r"\s+", " ", text)


def compact_rows(ws, min_row, max_col, maximum_height=54):
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=max_col):
        longest = max((len(str(cell.value or "")) for cell in row), default=0)
        ws.row_dimensions[row[0].row].height = min(maximum_height, 22 + 11 * min(3, longest // 45))
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_header(ws, cell_range):
    fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--records", required=True)
    parser.add_argument("--evidence", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--workspace-root", required=True)
    parser.add_argument("--umap-audit", required=True)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    workspace = resolved_e(args.workspace_root, "workspace root")
    records_path = within(resolved_e(args.records, "records"), workspace, "records")
    evidence_path = within(resolved_e(args.evidence, "evidence"), workspace, "evidence")
    umap_audit_path = within(resolved_e(args.umap_audit, "UMAP audit"), workspace, "UMAP audit")
    output = within(resolved_e(args.output, "workbook output"), workspace, "workbook output")
    if output.exists() and not args.force:
        raise FileExistsError(f"Refusing to overwrite existing workbook without --force: {output}")
    records = json.loads(records_path.read_text(encoding="utf-8"))
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    if not str(evidence.get("source_paths", {}).get("umap", "")).strip():
        raise ValueError("Formal workbook delivery is blocked because the evidence pack has no UMAP source")
    records = sorted(records, key=lambda record: cluster_sort_key(record.get("cluster_id", "")))
    inject_deterministic_evidence(records, evidence)
    sorted_clusters = sorted((str(cluster) for cluster in evidence["clusters"]), key=cluster_sort_key)
    umap_summary = validate_umap_audit(
        load_umap_audit(umap_audit_path), sorted_clusters, formal=True, records=records
    )
    apply_umap_audit(records, umap_summary)
    validate(records, sorted_clusters, evidence)
    label_normalization_changes = normalize_record_labels(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "注释结果"
    detail = wb.create_sheet("详细证据")
    source = wb.create_sheet("说明与数据来源")
    ws.append(MAIN_HEADERS)
    for record in records:
        ws.append([human_value(record.get(field, "")) for field in MAIN_FIELDS])
    detail.append(EVIDENCE_HEADERS)
    for record in records:
        detail.append([human_value(record.get(field, "")) for field in EVIDENCE_FIELDS])
    ws.freeze_panes = "A2"
    detail.freeze_panes = "A2"
    last_column = get_column_letter(len(MAIN_FIELDS))
    detail_last_column = get_column_letter(len(EVIDENCE_FIELDS))
    ws.sheet_view.showGridLines = False
    detail.sheet_view.showGridLines = False
    style_header(ws, f"A1:{last_column}1")
    style_header(detail, f"A1:{detail_last_column}1")
    ws.row_dimensions[1].height = 30
    detail.row_dimensions[1].height = 30
    thin = Side(style="thin", color="D9E2F3")
    for table, field_count in ((ws, len(MAIN_FIELDS)), (detail, len(EVIDENCE_FIELDS))):
        compact_rows(table, 2, field_count)
        for row in table.iter_rows(min_row=2, max_row=table.max_row, min_col=1, max_col=field_count):
            for cell in row:
                cell.border = Border(bottom=thin)
    set_widths(ws, [10, 22, 20, 16, 15, 15, 18, 28, 24, 24, 11, 11, 16, 18, 22, 42, 12, 36])
    set_widths(detail, [10, 18, 18, 24, 34, 20, 18, 16, 16, 20, 14, 18, 24, 14, 16, 28, 16, 14, 14, 28, 16, 18, 20, 14, 20, 14, 12, 16, 16, 18, 14, 18, 34, 16, 20, 20, 34, 24] + [16, 24, 36, 18, 28, 18, 24, 22, 18, 22, 48])
    quality_col = get_column_letter(MAIN_FIELDS.index("quality_score") + 1)
    mixed_col = get_column_letter(MAIN_FIELDS.index("mixed_or_doublet") + 1)
    review_col = get_column_letter(MAIN_FIELDS.index("manual_review") + 1)
    confidence_col = get_column_letter(MAIN_FIELDS.index("confidence") + 1)
    ws.conditional_formatting.add(f"{quality_col}2:{quality_col}{ws.max_row}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
    ws.conditional_formatting.add(f"{mixed_col}2:{mixed_col}{ws.max_row}", FormulaRule(formula=[f'{mixed_col}2="是"'], fill=PatternFill("solid", fgColor="F8CBAD")))
    ws.conditional_formatting.add(f"{review_col}2:{review_col}{ws.max_row}", FormulaRule(formula=[f'{review_col}2="是"'], fill=PatternFill("solid", fgColor="FFF2CC")))
    ws.conditional_formatting.add(f"{confidence_col}2:{confidence_col}{ws.max_row}", FormulaRule(formula=[f'{confidence_col}2="low"'], fill=PatternFill("solid", fgColor="FCE4D6")))

    metadata = evidence.get("confirmed_metadata", {})
    paths = evidence.get("source_paths", {})
    source_rows = [
        ["项目", f"{metadata.get('species', '')} {metadata.get('tissue', '')} {metadata.get('parent_population', '')} 亚群注释"],
        ["物种", metadata.get("species", "")], ["组织/解剖来源", metadata.get("tissue", "")],
        ["注释层级", metadata.get("annotation_level", "")], ["父群", metadata.get("parent_population", "")],
        ["父群类型", metadata.get("parent_kind", "")], ["父群解释规则", metadata.get("interpretation_rule", "")],
        ["平均表达源文件", Path(paths.get("cell_avg_exp", "")).name],
        ["Marker源文件", Path(paths.get("marker_table", "")).name],
        ["完整表达占比源文件", Path(paths.get("expression_ratio_table", "")).name or "未提供"],
        ["跨物种基因映射", Path(paths.get("gene_map", "")).name or "未提供"],
        ["单细胞验证证据", Path(paths.get("cell_evidence", "")).name or "未提供"],
        ["输入规模", f"{evidence['average_shape'][0]:,} genes × {evidence['average_shape'][1]} clusters"],
        ["平均表达读取", evidence.get("average_reader", "")],
        ["命名模式", "优先使用通行短名称；身份、发育阶段和状态分别记录"],
        ["候选策略", "开放生成候选，使用组合 Marker、冲突证据和数据集内参照严格验证"],
        ["跨物种策略", "缺少物种专属面板时使用同源基因与保守表达程序推断并降低置信度"],
        ["判定原则", "多基因一致性、表达占比、组间特异性和同数据集阳性参照共同判定"],
        ["确定性证据配置", evidence.get("annotation_evidence_policy", {}).get("config_version", "")],
        ["重点复核", "、".join(str(r["cluster_id"]) for r in records if r["manual_review"]) or "无"],
    ]
    source_rows.extend([
        ["UMAP源文件", Path(paths.get("umap", "")).name],
        ["UMAP结构化审核", umap_audit_path.name],
    ])
    for row in source_rows:
        source.append(row)
    source.sheet_view.showGridLines = False
    set_widths(source, [24, 72])
    for row in source.iter_rows():
        row[0].fill = PatternFill("solid", fgColor="D9EAF7")
        row[0].font = Font(bold=True, color="17365D")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        source.row_dimensions[row[0].row].height = 28
    style_header(source, "A1:B1")

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    check = load_workbook(output, read_only=False, data_only=False)
    expected_sheets = ["注释结果", "详细证据", "说明与数据来源"]
    if check.sheetnames != expected_sheets:
        raise RuntimeError(f"Workbook sheet verification failed: {check.sheetnames}")
    if check["注释结果"].max_row != len(records) + 1 or check["详细证据"].max_row != len(records) + 1:
        raise RuntimeError("Workbook row-count verification failed")
    result_order = [str(check["注释结果"].cell(row, 1).value) for row in range(2, len(records) + 2)]
    detail_order = [str(check["详细证据"].cell(row, 1).value) for row in range(2, len(records) + 2)]
    if result_order != sorted_clusters or detail_order != sorted_clusters:
        raise RuntimeError(
            f"Workbook cluster-order verification failed: expected={sorted_clusters}, "
            f"result={result_order}, detail={detail_order}"
        )
    qa = {
        "status": "pass", "workbook": str(output), "sheets": check.sheetnames,
        "record_count": len(records), "review_clusters": [str(r["cluster_id"]) for r in records if r["manual_review"]],
        "mixed_or_doublet_clusters": [str(r["cluster_id"]) for r in records if r["mixed_or_doublet"]],
        "formula_cells": 0, "source_files_unchanged": True,
        "visual_qa": "compact_human_record_layout",
        "auto_filter_enabled": False,
        "main_fields": MAIN_FIELDS,
        "evidence_fields": EVIDENCE_FIELDS,
        "cluster_order": sorted_clusters,
        "cluster_order_policy": "numeric_ascending_then_natural_alphanumeric",
        "hierarchy_depth_conflicts": hierarchy_depth_conflicts(records),
        "hierarchy_depth_policy": "Do not mix an ontology ancestor with any of its descendants in one final subcluster table.",
        "developmental_stage_by_cluster": {
            str(record["cluster_id"]): str(record.get("developmental_stage", "")) for record in records
        },
        "tissue_context_review_clusters": [
            str(record["cluster_id"]) for record in records if record.get("tissue_context_review")
        ],
        "naming_grammars": sorted({str(r["naming_grammar"]) for r in records}),
        "final_label_policy": "letters_digits_cjk_underscore_only_and_no_redundant_t_suffix",
        "final_labels_normalized": True,
        "label_normalization_changes": label_normalization_changes,
        "label_basis_counts": {
            "canonical_subtype": sum(r["label_basis"] == "canonical_subtype" for r in records),
            "validated_external_candidate": sum(r["label_basis"] == "validated_external_candidate" for r in records),
            "top_marker_fallback": sum(r["label_basis"] == "top_marker_fallback" for r in records),
            "feature_gene_fallback": sum(r["label_basis"] == "feature_gene_fallback" for r in records),
        },
        "naming_marker_policy": evidence.get("naming_marker_policy", {}),
        "naming_marker_by_cluster": {
            str(cluster): (
                evidence["cluster_profiles"][str(cluster)].get("naming_top_marker") or {}
            ).get("gene", "")
            for cluster in sorted_clusters
        },
        "selected_label_marker_by_cluster": {
            str(record["cluster_id"]): str(record["top_marker_gene"]) for record in records
        },
        "contextually_excluded_naming_markers_by_cluster": {
            str(record["cluster_id"]): record["contextually_excluded_naming_markers"] for record in records
        },
        "raw_top_marker_by_cluster": {
            str(cluster): (
                evidence["cluster_profiles"][str(cluster)].get("raw_top_marker") or {}
            ).get("gene", "")
            for cluster in sorted_clusters
        },
        "excluded_naming_markers_by_cluster": {
            str(cluster): evidence["cluster_profiles"][str(cluster)].get("excluded_naming_markers", [])
            for cluster in sorted_clusters
        },
        "annotation_evidence_policy": evidence.get("annotation_evidence_policy", {}),
        "deterministic_risk_by_cluster": {str(record["cluster_id"]): record["risk_level"] for record in records},
        "deterministic_action_by_cluster": {str(record["cluster_id"]): record["recommended_action"] for record in records},
        "umap_reviewed_clusters": umap_summary["reviewed_clusters"],
        "umap_conflict_clusters": umap_summary["conflict_clusters"],
        "umap_research_pending_clusters": umap_summary["research_pending_clusters"],
        "umap_all_clusters_reviewed": umap_summary["all_clusters_reviewed"],
    }
    qa_path = output.with_suffix(".qa.json")
    qa_path.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    qa["case_registry"] = register_shared_case(workspace, records_path, evidence_path, qa_path, output)
    qa["delivery"] = deliver_final_workbook(workspace, output, evidence)
    qa_path.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(qa, ensure_ascii=False))


if __name__ == "__main__":
    main()



