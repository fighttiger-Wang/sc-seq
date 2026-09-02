#!/usr/bin/env python3
"""Shared qualitative workbook contract for major and subcluster annotation."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import platform
import re
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from qualitative_umap_resolution import apply_identity_resolution, normalize_relation, validate_identity_resolution


GATE_VALUES = {"通过", "不通过", "未确定", "不适用"}
PLOT_LABEL = re.compile(r"^[A-Za-z0-9_]+$")
INVALID_LABEL = re.compile(r"[^A-Za-z0-9_]+")

PLOT_FIELDS = ["cluster_id", "celltype_en"]
PLOT_HEADERS = ["Cluster", "Celltype_EN"]

RESULT_FIELDS = [
    "cluster_id", "celltype_cn", "celltype_en", "broad_type",
    "developmental_stage", "state", "disease_role", "key_markers",
    "candidate_labels", "umap_summary", "boundary_flags",
    "possible_components", "rationale", "validation_advice", "handling_advice",
]
RESULT_HEADERS = [
    "Cluster", "中文名称", "Celltype_EN", "细胞谱系", "发育/成熟阶段",
    "细胞状态", "组织/疾病相关角色", "关键 Marker", "主要竞争候选",
    "UMAP 判断摘要", "异常/边界标记", "可能组成", "判定摘要", "验证建议",
    "下游处理建议",
]

EVIDENCE_FIELDS = [
    "cluster_id", "celltype_cn", "celltype_en", "parent_context",
    "primary_program", "competing_programs", "supporting_marker_evidence",
    "conflicting_marker_evidence", "missing_markers", "identity_anchor_gate",
    "parent_lineage_gate", "sibling_competition_gate", "exclusion_gate",
    "off_parent_gate", "off_parent_audit", "developmental_program",
    "state_program", "state_gate", "umap_gate", "umap_summary",
    "cross_island_audit", "mixed_doublet_gate", "mixed_doublet_explanation",
    "rationale", "evidence_gaps", "validation_advice", "handling_advice",
]
EVIDENCE_HEADERS = [
    "Cluster", "中文名称", "Celltype_EN", "父群/谱系背景", "主要身份程序",
    "竞争身份程序", "支持 Marker 证据", "冲突 Marker 证据", "缺失 Marker",
    "身份锚点门控", "父谱系门控", "同级竞争门控", "排除证据门控",
    "离群/跨谱系门控", "离群/跨谱系审计", "发育/成熟程序", "状态程序",
    "状态门控", "UMAP 门控", "UMAP 判断摘要", "跨岛一致性审计",
    "混合/双细胞门控", "混合/双细胞解释", "判定依据", "证据缺口",
    "验证建议", "下游处理建议",
]

LITERATURE_HEADERS = ["细胞类型", "文献", "经典鉴定 Marker", "本次鉴定使用的 Marker"]
SOURCE_HEADERS = ["项目", "内容"]


def cluster_sort_key(value):
    text = str(value).strip()
    try:
        number = Decimal(text)
        if number.is_finite():
            return (0, number, text)
    except InvalidOperation:
        pass
    natural = tuple(
        (0, int(part)) if part.isdigit() else (1, part.lower())
        for part in re.split(r"(\d+)", text)
    )
    return (1, natural, text)


def normalize_final_label(value):
    text = INVALID_LABEL.sub("_", str(value or "").strip())
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        raise ValueError(f"Final plotting label is empty after normalization: {value!r}")
    return text


def _as_list(value):
    if value in (None, ""):
        return []
    if isinstance(value, (list, tuple, set)):
        return list(value)
    if isinstance(value, dict):
        return [value]
    text = str(value).strip()
    if text.startswith(("[", "{")):
        try:
            parsed = json.loads(text)
            return _as_list(parsed)
        except json.JSONDecodeError:
            pass
    return [item.strip() for item in re.split(r"[;,；、]", text) if item.strip()]


def human_value(value):
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, dict):
        return "；".join(
            f"{key}：{human_value(item)}" for key, item in value.items()
            if human_value(item) not in (None, "")
        )
    if isinstance(value, (list, tuple, set)):
        return "；".join(str(human_value(item)) for item in value if human_value(item) not in (None, ""))
    return re.sub(r"\s+", " ", str(value).strip())


def _metric_value(item, *names):
    for name in names:
        if name in item and item[name] not in (None, ""):
            return item[name]
    return None


def _number_text(value, percent=False):
    if value in (None, ""):
        return "NA"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if percent:
        if abs(number) <= 1.0:
            number *= 100.0
        return f"{number:.2f}%"
    return f"{number:.2f}"


def format_marker_metric(item):
    if isinstance(item, str):
        return item.strip()
    gene = str(_metric_value(item, "gene", "Gene", "GeneName") or "").strip()
    if not gene:
        return human_value(item)
    mean = _metric_value(item, "mean_expr", "mean", "Target_Cluster_mean", "avg_expr")
    ratio = _metric_value(item, "expr_ratio", "ratio", "detection_ratio", "p_in", "pct.1", "pct1")
    log2fc = _metric_value(item, "log2FC", "avg_log2FC", "avg_logFC")
    pct1 = _metric_value(item, "pct.1", "pct1", "p_in")
    pct2 = _metric_value(item, "pct.2", "pct2", "p_background", "background")
    return (
        f"{gene}(mean={_number_text(mean)}, ratio={_number_text(ratio, True)}, "
        f"log2FC={_number_text(log2fc)}, pct.1={_number_text(pct1, True)}, "
        f"pct.2={_number_text(pct2, True)})"
    )


def _profile_markers(evidence, cluster):
    profile = evidence.get("cluster_profiles", {}).get(str(cluster), {})
    return profile.get("top_markers", []) or profile.get("top_informative_markers", []) or []


def _marker_lookup(evidence, cluster):
    lookup = {}
    for item in _profile_markers(evidence, cluster):
        if isinstance(item, dict):
            gene = str(_metric_value(item, "gene", "Gene", "GeneName") or "").upper()
            if gene:
                lookup[gene] = item
    return lookup


def _marker_evidence(value, evidence, cluster):
    lookup = _marker_lookup(evidence, cluster)
    rendered = []
    for item in _as_list(value):
        if isinstance(item, dict):
            rendered.append(format_marker_metric(item))
            continue
        gene = str(item).strip()
        rendered.append(format_marker_metric(lookup.get(gene.upper(), gene)))
    return "；".join(item for item in rendered if item)


def _gate(value, default="未确定"):
    if isinstance(value, dict):
        value = value.get("status", value.get("result", value.get("gate", "")))
    if isinstance(value, bool):
        return "通过" if value else "不通过"
    text = str(value or "").strip()
    aliases = {
        "pass": "通过", "passed": "通过", "true": "通过",
        "fail": "不通过", "failed": "不通过", "false": "不通过",
        "unknown": "未确定", "uncertain": "未确定", "not_assessed": "未确定",
        "na": "不适用", "n/a": "不适用", "not_applicable": "不适用",
    }
    result = aliases.get(text.lower(), text)
    return result if result in GATE_VALUES else default


def _decision(evidence, cluster):
    return evidence.get("qualitative_annotation_evidence", {}).get(str(cluster), {})


def _umap_entry(umap_audit, cluster):
    return (umap_audit or {}).get("clusters", {}).get(str(cluster), {})


def _first(record, decision, *keys, default=""):
    for source in (record, decision):
        for key in keys:
            value = source.get(key)
            if value not in (None, "", [], {}):
                return value
    return default


def _qualitative_gates(record, decision, umap):
    gates = dict(decision.get("qualitative_gates", {}))
    identity = _gate(gates.get("identity_anchor"))
    if identity == "未确定" and decision.get("stable_id"):
        identity = "通过"
    parent = _gate(gates.get("parent_lineage"), "不适用")
    siblings = _gate(gates.get("sibling_competition"))
    exclusions = _gate(gates.get("exclusion"))
    off_parent = _gate(gates.get("off_parent"), "不适用")
    state = _gate(gates.get("state_program"), "不适用")
    umap_gate = _gate(gates.get("umap"), "未确定")
    relation = normalize_relation(umap.get("marker_umap_relation", ""))
    if relation == "concordant":
        umap_gate = "通过"
    elif relation == "conflict":
        umap_gate = "不通过"
    mixed = _gate(gates.get("mixed_doublet"), "不适用")
    if any(bool(_first(record, decision, key)) for key in ("mixed_population", "suspected_doublet", "mixed_evidence")):
        mixed = "未确定" if not decision.get("mixed_population") else "通过"
    return identity, parent, siblings, exclusions, off_parent, state, umap_gate, mixed


def normalize_records(records, evidence, umap_audit=None):
    normalized = []
    for raw in records:
        record = dict(raw)
        cluster = str(record.get("cluster_id", "")).strip()
        decision = _decision(evidence, cluster)
        umap = _umap_entry(umap_audit, cluster)
        if umap:
            apply_identity_resolution(record, umap)
        stable = _first(record, decision, "stable_id", "celltype_en", "display_label")
        celltype_en = normalize_final_label(_first(record, decision, "celltype_en", "display_label", default=stable))
        record.update({
            "cluster_id": cluster,
            "celltype_en": celltype_en,
            "celltype_cn": _first(record, decision, "celltype_cn", default=celltype_en),
            "stable_id": normalize_final_label(stable or celltype_en),
            "broad_type": _first(record, decision, "broad_type", "primary_major_label", "expected_parent_id"),
            "developmental_stage": _first(record, decision, "developmental_stage"),
            "state": _first(record, decision, "state", "primary_state"),
            "disease_role": _first(record, decision, "disease_role"),
            "key_markers": human_value(_first(record, decision, "key_markers", "supporting_markers")),
            "candidate_labels": human_value(_first(record, decision, "candidate_labels", "competing_programs", "possible_components")),
            "possible_components": human_value(_first(record, decision, "possible_components")),
            "rationale": human_value(_first(record, decision, "rationale", "decision_rationale")),
            "validation_advice": human_value(_first(record, decision, "validation_advice", "review_action", "recommended_action")),
            "handling_advice": human_value(_first(record, decision, "handling_advice", "downstream_handling", "review_action", "recommended_action")),
            "umap_summary": human_value(_first(record, umap, "umap_summary", "topology_summary", "review_action")),
            "boundary_flags": human_value(_first(record, decision, "boundary_flags", "abnormality_flags", "interpretation_flags", "mixture_type")),
            "parent_context": human_value(_first(record, decision, "parent_context", "parent_path", "expected_parent_id")),
            "primary_program": human_value(_first(record, decision, "primary_program", "primary_evidence_label", "stable_id", default=celltype_en)),
            "competing_programs": human_value(_first(record, decision, "competing_programs", "candidate_labels", "possible_components")),
            "supporting_marker_evidence": _marker_evidence(
                _first(record, decision, "supporting_marker_evidence", "supporting_markers"), evidence, cluster
            ),
            "conflicting_marker_evidence": _marker_evidence(
                _first(record, decision, "conflicting_marker_evidence", "conflicting_markers", "negative_marker_conflict"), evidence, cluster
            ),
            "missing_markers": human_value(_first(record, decision, "missing_markers", "evidence_gaps")),
            "off_parent_audit": human_value(_first(record, decision, "off_parent_audit", "off_parent_candidate", "tissue_context_review")),
            "developmental_program": human_value(_first(record, decision, "developmental_program", "developmental_stage")),
            "state_program": human_value(_first(record, decision, "state_program", "state_list")),
            "cross_island_audit": human_value(_first(record, umap, "cross_island_audit", "same_label_topology", "separation_explanation", "separation_evidence")),
            "mixed_doublet_explanation": human_value(_first(record, decision, "mixed_doublet_explanation", "mixture_type", "possible_components")),
            "evidence_gaps": human_value(_first(record, decision, "evidence_gaps", "missing_markers")),
        })
        gates = _qualitative_gates(record, decision, umap)
        for field, value in zip(
            ("identity_anchor_gate", "parent_lineage_gate", "sibling_competition_gate",
             "exclusion_gate", "off_parent_gate", "state_gate", "umap_gate", "mixed_doublet_gate"),
            gates,
        ):
            record[field] = value
        record["_decision"] = decision
        record["_umap"] = umap
        normalized.append(record)
    return sorted(normalized, key=lambda item: cluster_sort_key(item["cluster_id"]))


def validate(records, clusters, evidence, umap_audit=None, annotation_level="major"):
    errors = []
    expected = sorted((str(item) for item in clusters), key=cluster_sort_key)
    observed = [str(item.get("cluster_id", "")) for item in records]
    if observed != expected:
        errors.append(f"Cluster order/membership mismatch: expected={expected}, records={observed}")
    if len(set(observed)) != len(observed):
        errors.append("Duplicate cluster IDs are not allowed")
    umap_source = str(evidence.get("source_paths", {}).get("umap", "")).strip()
    if umap_source:
        if not isinstance(umap_audit, dict):
            errors.append("Formal delivery has a UMAP source but lacks a cluster-complete UMAP audit")
        else:
            audit_clusters = umap_audit.get("clusters", {})
            missing_audit = sorted(set(expected) - set(map(str, audit_clusters)), key=cluster_sort_key)
            if missing_audit:
                errors.append(f"UMAP audit lacks clusters: {missing_audit}")
            for cluster in expected:
                audit = audit_clusters.get(str(cluster), {})
                if not audit.get("reviewed"):
                    errors.append(f"Cluster {cluster} UMAP audit is not marked reviewed")
                if not str(audit.get("topology_summary", "")).strip():
                    errors.append(f"Cluster {cluster} UMAP audit lacks topology_summary")
                relation = normalize_relation(audit.get("marker_umap_relation", ""))
                audit["marker_umap_relation"] = relation
                if relation not in {"concordant", "conflict", "indeterminate"}:
                    errors.append(f"Cluster {cluster} invalid marker_umap_relation: {relation}")
                matching = next((record for record in records if str(record.get("cluster_id", "")) == cluster), {})
                errors.extend(validate_identity_resolution(
                    audit, matching, evidence=evidence, formal=annotation_level == "subcluster"
                ))
                if relation == "conflict":
                    if audit.get("research_status") not in {"resolved", "reused"}:
                        errors.append(f"Cluster {cluster} unresolved marker/UMAP conflict blocks formal delivery")
                    if audit.get("conflict_resolution_basis") == "literature_only":
                        errors.append(f"Cluster {cluster} sample-specific UMAP conflict cannot be resolved by literature alone")
    for record in records:
        cluster = record.get("cluster_id", "")
        for field in ("celltype_cn", "celltype_en", "rationale", "validation_advice", "handling_advice"):
            if not str(record.get(field, "")).strip():
                errors.append(f"Cluster {cluster} lacks required qualitative field: {field}")
        if not PLOT_LABEL.fullmatch(str(record.get("celltype_en", ""))):
            errors.append(f"Cluster {cluster} Celltype_EN violates [A-Za-z0-9_]+")
        for field in ("identity_anchor_gate", "parent_lineage_gate", "sibling_competition_gate",
                      "exclusion_gate", "off_parent_gate", "state_gate", "umap_gate", "mixed_doublet_gate"):
            if record.get(field) not in GATE_VALUES:
                errors.append(f"Cluster {cluster} invalid qualitative gate {field}: {record.get(field)}")
        if annotation_level == "subcluster":
            parent = str(evidence.get("confirmed_metadata", {}).get("parent_population", "")).strip()
            if parent and record.get("celltype_en") == normalize_final_label(parent):
                errors.append(f"Cluster {cluster} retreats to the supplied parent instead of a sibling/leaf identity")
    if errors:
        raise ValueError("\n".join(errors))


def inject_qualitative_evidence(records, evidence):
    """Enrich records with qualitative decision fields only."""
    normalized = normalize_records(records, evidence)
    records[:] = normalized
    return records


# Temporary import compatibility for historical tests; no deterministic
# evidence is read or generated by this alias.
inject_deterministic_evidence = inject_qualitative_evidence


def _warning(record):
    decision = record.get("_decision", {})
    if record.get("stable_id") == "Multi_cell" or record.get("celltype_en") == "Multi_cell":
        return True
    for key in (
        "suspected_doublet", "low_quality", "debris", "background_interference",
        "off_parent_detected", "off_parent_reassignment", "mixed_population",
        "mixed_evidence", "lineage_boundary", "abnormal_state",
    ):
        if bool(_first(record, decision, key)):
            return True
    state_program = decision.get("state_program", [])
    if isinstance(state_program, list):
        return any(
            isinstance(item, dict) and int(item.get("marker_count", 0) or 0) >= 2
            for item in state_program
        )
    return False


def _citation_text(source):
    if isinstance(source, str):
        return source.strip(), ""
    author = source.get("first_author") or source.get("author") or source.get("authors") or ""
    year = source.get("year") or ""
    journal = source.get("journal") or ""
    title = source.get("title") or ""
    pmid = source.get("pmid") or ""
    doi = source.get("doi") or ""
    locator = pmid or doi
    text = ", ".join(str(item) for item in (author, year, journal, title, locator) if item)
    url = source.get("url") or source.get("link") or ""
    if not url and doi:
        url = f"https://doi.org/{str(doi).replace('DOI:', '').strip()}"
    if not url and pmid:
        digits = re.sub(r"\D", "", str(pmid))
        if digits:
            url = f"https://pubmed.ncbi.nlm.nih.gov/{digits}/"
    return text, url


def _literature_rows(records, evidence):
    grouped = {}
    for record in records:
        label = record["celltype_en"]
        entry = grouped.setdefault(label, {"sources": [], "classic": [], "used": []})
        entry["classic"].extend(_as_list(record.get("classic_markers")))
        entry["used"].extend(_as_list(record.get("key_markers")))
        sources = record.get("literature_details") or record.get("literature_source")
        entry["sources"].extend(_as_list(sources))
        if not entry["sources"]:
            decision = record.get("_decision", {})
            source_ids = _as_list(decision.get("marker_panel_evidence_ids"))
            registry = evidence.get("annotation_evidence_policy", {}).get("evidence_source_registry", {})
            entry["sources"].extend(registry.get(item, item) for item in source_ids)
    rows = []
    for label, entry in grouped.items():
        sources = entry["sources"] or ["未提供结构化文献；正式交付前需补充"]
        classic = "；".join(dict.fromkeys(str(item) for item in entry["classic"] if str(item).strip()))
        used = "；".join(dict.fromkeys(str(item) for item in entry["used"] if str(item).strip()))
        for source in sources:
            text, url = _citation_text(source)
            rows.append((label, text, classic, used, url))
    return rows


def _source_rows(evidence, annotation_level, records, skill_name, skill_version):
    metadata = evidence.get("confirmed_metadata", {})
    paths = evidence.get("source_paths", {})
    policy = evidence.get("annotation_evidence_policy", {})
    filenames = [Path(str(value)).name for value in paths.values() if str(value).strip()]
    constraints = metadata.get("annotation_constraints") or policy.get("user_constraints") or "无"
    return [
        ["物种", metadata.get("species", "")],
        ["组织/器官", metadata.get("tissue", "")],
        ["父群", metadata.get("parent_population", "") or "不适用"],
        ["注释模式", annotation_level],
        ["输入文件", "；".join(dict.fromkeys(filenames))],
        ["指标定义", "mean_expr=平均表达；expr_ratio/pct.1=目标群检出比例；pct.2=背景群检出比例；log2FC=差异表达倍数"],
        ["Skill 版本", f"{skill_name} {skill_version}"],
        ["证据核心/配置版本", f"core {policy.get('core_version', '')}；config {policy.get('config_version', '')}"],
        ["生成时间", datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")],
        ["限定条件", human_value(constraints)],
        ["排序规则", "所有含 Cluster 的表均按数值升序，再按自然字母数字顺序"],
        ["红色填充规则", "仅中文名称列；用于显著状态程序、Multi_cell、疑似双细胞、低质量/碎片/背景干扰及谱系边界"],
        ["证据解释", "不使用评分或置信度；原始数值仅作为单个 Marker 的证据，身份由生物学门控和专家推理决定"],
        ["局限性", "Cluster 汇总证据不能证明同一细胞共表达，也不能确认 doublet；UMAP 仅作一致性审计"],
        ["数据处理声明", "注释过程未自动删除、过滤、合并或修改细胞及原始数据"],
        ["Cluster 数量", len(records)],
    ]


def _style_sheet(ws, widths, freeze, row_height=24):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = None
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=False)
    ws.row_dimensions[1].height = 26
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = row_height
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False, shrink_to_fit=False)
            cell.border = Border(bottom=thin)
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def build_workbook(records, evidence, output, annotation_level, skill_name, skill_version, umap_audit=None):
    clusters = sorted((str(item) for item in evidence.get("clusters", [])), key=cluster_sort_key)
    records = normalize_records(records, evidence, umap_audit)
    validate(records, clusters, evidence, umap_audit, annotation_level)

    wb = Workbook()
    plot = wb.active
    plot.title = "绘图列表"
    result = wb.create_sheet("注释结果")
    detail = wb.create_sheet("详细证据")
    literature = wb.create_sheet("细胞类型与文献")
    source = wb.create_sheet("说明与数据来源")

    plot.append(PLOT_HEADERS)
    result.append(RESULT_HEADERS)
    detail.append(EVIDENCE_HEADERS)
    literature.append(LITERATURE_HEADERS)
    source.append(SOURCE_HEADERS)

    for record in records:
        plot.append([human_value(record[field]) for field in PLOT_FIELDS])
        result.append([human_value(record.get(field, "")) for field in RESULT_FIELDS])
        detail.append([human_value(record.get(field, "")) for field in EVIDENCE_FIELDS])

    for label, citation, classic, used, url in _literature_rows(records, evidence):
        literature.append([label, citation, classic, used])
        if url:
            literature.cell(literature.max_row, 2).hyperlink = url
            literature.cell(literature.max_row, 2).style = "Hyperlink"
    for row in _source_rows(evidence, annotation_level, records, skill_name, skill_version):
        source.append(row)

    _style_sheet(plot, [12, 26], "A2", 22)
    _style_sheet(result, [10, 22, 22, 18, 18, 18, 24, 36, 28, 36, 28, 28, 44, 44, 44], "D2", 24)
    _style_sheet(detail, [10, 22, 22, 28, 30, 30, 54, 48, 36, 16, 16, 16, 16, 16, 40, 28, 34, 16, 16, 40, 36, 18, 40, 54, 40, 48, 48], "D2", 24)
    _style_sheet(literature, [24, 72, 44, 44], "B2", 24)
    _style_sheet(source, [24, 88], "A2", 24)

    red_fill = PatternFill("solid", fgColor="FFF8696B")
    for row_index, record in enumerate(records, 2):
        if _warning(record):
            result.cell(row_index, RESULT_FIELDS.index("celltype_cn") + 1).fill = red_fill

    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    check = load_workbook(output, read_only=False, data_only=False)
    expected_sheets = ["绘图列表", "注释结果", "详细证据", "细胞类型与文献", "说明与数据来源"]
    if check.sheetnames != expected_sheets:
        raise RuntimeError(f"Workbook sheet verification failed: {check.sheetnames}")
    for sheet, headers in (("绘图列表", PLOT_HEADERS), ("注释结果", RESULT_HEADERS), ("详细证据", EVIDENCE_HEADERS), ("细胞类型与文献", LITERATURE_HEADERS), ("说明与数据来源", SOURCE_HEADERS)):
        if [cell.value for cell in check[sheet][1]] != headers:
            raise RuntimeError(f"{sheet} header verification failed")
        if check[sheet].auto_filter.ref is not None:
            raise RuntimeError(f"{sheet} must not enable autofilter")
    for sheet in ("绘图列表", "注释结果", "详细证据"):
        order = [str(check[sheet].cell(row, 1).value) for row in range(2, len(records) + 2)]
        if order != clusters:
            raise RuntimeError(f"{sheet} cluster order mismatch: {order} != {clusters}")
    deprecated = sorted({key for record in records for key in record if key in {
        "confidence", "quality_score", "primary_evidence_score", "runner_up_evidence_score",
        "score_margin", "rival_lineage_score", "ranked_identity_evidence"
    }})
    digest = hashlib.sha256(output.read_bytes()).hexdigest()
    return {
        "status": "pass", "workbook": str(output), "sheets": expected_sheets,
        "workbook_sha256": digest,
        "record_count": len(records), "cluster_order": clusters,
        "cluster_order_policy": "numeric_ascending_then_natural_alphanumeric",
        "qualitative_gate_values": sorted(GATE_VALUES), "aggregate_scores_exported": False,
        "confidence_exported": False, "deprecated_input_fields_ignored": deprecated,
        "red_fill_column": "中文名称", "auto_filter_enabled": False,
        "fixed_row_height": True, "wrap_text": False, "shrink_to_fit": False,
        "source_files_unchanged": True,
    }


def resolved_e(path, role):
    resolved = Path(path).resolve()
    if platform.system() == "Windows" and os.path.splitdrive(str(resolved))[0].upper() != "E:":
        raise ValueError(f"{role} must be on E: {resolved}")
    return resolved


def within(path, root, role):
    path, root = Path(path).resolve(), Path(root).resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"{role} must be inside workspace root {root}: {path}") from exc
    return path


def cli_main(annotation_level, skill_name, skill_version):
    parser = argparse.ArgumentParser()
    parser.add_argument("--records", required=True)
    parser.add_argument("--evidence", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--workspace-root", required=True)
    parser.add_argument("--umap-audit")
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()
    workspace = resolved_e(args.workspace_root, "workspace root")
    records_path = within(resolved_e(args.records, "records"), workspace, "records")
    evidence_path = within(resolved_e(args.evidence, "evidence"), workspace, "evidence")
    output = within(resolved_e(args.output, "workbook output"), workspace, "workbook output")
    if output.exists() and not args.force:
        raise FileExistsError(f"Refusing to overwrite existing workbook without --force: {output}")
    umap_audit = None
    if args.umap_audit:
        audit_path = within(resolved_e(args.umap_audit, "UMAP audit"), workspace, "UMAP audit")
        umap_audit = json.loads(audit_path.read_text(encoding="utf-8"))
    records = json.loads(records_path.read_text(encoding="utf-8"))
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    qa = build_workbook(records, evidence, output, annotation_level, skill_name, skill_version, umap_audit)
    output.with_suffix(".qa.json").write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(qa, ensure_ascii=False))
