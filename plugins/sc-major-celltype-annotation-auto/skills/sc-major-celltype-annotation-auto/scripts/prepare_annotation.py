#!/usr/bin/env python3
"""One-command preflight for paired-Excel cluster annotation."""

import argparse
import json
import os
import platform
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from annotation_evidence_core import enrich_evidence
from inspect_annotation_inputs import build_evidence


STATE_PARENT_TOKENS = ("dividing", "cycling", "prolifer", "activated", "stress", "interferon", "hypoxi", "增殖", "活化", "应激")


def drive(path):
    return os.path.splitdrive(str(Path(path).resolve()))[0].upper()


def assert_e_drive(path, role):
    resolved = Path(path).resolve()
    if platform.system() == "Windows" and drive(resolved) != "E:":
        raise ValueError(f"{role} must be on E: under the workspace policy: {resolved}")
    return resolved


def assert_within(path, root, role):
    path, root = Path(path).resolve(), Path(root).resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"{role} must stay inside workspace root {root}: {path}") from exc
    return path


def has_reparse_component(path, stop):
    current, stop = Path(path).resolve(), Path(stop).resolve()
    while True:
        if current.exists() and (current.is_symlink() or (hasattr(os.path, "isjunction") and os.path.isjunction(current))):
            return str(current)
        if current == stop or current.parent == current:
            return None
        current = current.parent


def infer_parent_kind(parent, requested):
    if requested != "auto":
        return requested
    lower = parent.lower()
    return "state" if any(token in lower for token in STATE_PARENT_TOKENS) else "lineage"


def load_project_prior(path):
    if path is None:
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip().lower() for value in next(rows)]

    def column(*aliases):
        for alias in aliases:
            if alias.lower() in headers:
                return headers.index(alias.lower())
        raise ValueError(f"Project prior lacks one of the required columns: {aliases}")

    cluster_i = column("cluster", "seurat_clusters", "target_cluster")
    label_i = column("cells", "celltype", "celltype_en", "label", "annotation")
    prior = {}
    for row in rows:
        if row[cluster_i] is None or row[label_i] is None:
            continue
        cluster, label = str(row[cluster_i]).strip(), str(row[label_i]).strip()
        if cluster in prior and prior[cluster] != label:
            raise ValueError(f"Project prior contains conflicting labels for cluster {cluster}")
        prior[cluster] = label
    return prior


def load_sample_context(path):
    if path is None:
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Sample context must be a JSON object")
    return data


def load_annotation_constraints(path, exclude_labels, conflict_markers):
    constraints = {}
    if path is not None:
        constraints = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(constraints, dict):
            raise ValueError("Annotation constraints must be a JSON object")
    constraints = dict(constraints)
    constraints["exclude_labels"] = list(constraints.get("exclude_labels", [])) + list(exclude_labels or [])
    constraints["conflict_markers"] = list(
        constraints.get("conflict_markers", constraints.get("exclude_markers", []))
    ) + list(conflict_markers or [])
    return constraints


def template_record(cluster):
    return {
        "cluster_id": cluster, "celltype_cn": "", "celltype_en": "",
        "stable_id": "", "parent_path": [], "tissue_module": [],
        "disease_role": [], "state_list": [], "primary_state": "",
        "cross_species_inference": False, "panel_species": "", "marker_panel_evidence_ids": [],
        "label_basis": "", "canonical_subtype": "", "top_marker_gene": "",
        "literature_source": "", "naming_grammar": "",
        "contextually_excluded_naming_markers": [],
        "user_constraint_audit": {},
        "broad_type": "", "fine_type": "", "state": "",
        "supporting_markers": "", "conflicting_markers": "",
        "candidate_labels": "", "confidence": "", "quality_score": None,
        "mixed_or_doublet": False, "mixture_type": "none", "possible_components": "",
        "mixed_evidence": False, "review_in_subcluster": False,
        "rationale": "", "manual_review": True, "review_action": "",
    }


def evidence_digest(evidence):
    """Return compact model-facing evidence; retain the full pack for targeted review."""
    profiles = {}
    for cluster in evidence["clusters"]:
        profile = evidence["cluster_profiles"][str(cluster)]
        profiles[str(cluster)] = {
            "marker_count": profile["marker_count"],
            "top_markers": profile["top_markers"][:30],
            "raw_top_marker": profile["raw_top_marker"],
            "naming_top_marker": profile["naming_top_marker"],
            "excluded_naming_markers": profile["excluded_naming_markers"],
            "top_informative_markers": profile["top_informative_markers"][:15],
            "signature_scores": profile["signature_scores"],
            "signature_marker_support": profile.get("signature_marker_support", {}),
            "ranked_lineage_signatures": profile["ranked_lineage_signatures"],
            "qc_state_fraction_top50": profile["qc_state_fraction_top50"],
            "alerts": profile["alerts"],
            "deterministic_evidence": evidence.get("deterministic_annotation_evidence", {}).get(str(cluster), {}),
        }
    return {
        "schema_version": evidence["schema_version"],
        "average_reader": evidence["average_reader"],
        "average_gene_header": evidence.get("average_gene_header", "GeneName"),
        "average_gene_header_normalized_to": evidence.get("average_gene_header_normalized_to", "GeneName"),
        "average_shape": evidence["average_shape"],
        "average_matrix_semantics": evidence.get("average_matrix_semantics", {}),
        "clusters": evidence["clusters"],
        "cluster_profiles": profiles,
        "canonical_expression_by_gene": evidence["canonical_expression_by_gene"],
        "confirmed_metadata": evidence["confirmed_metadata"],
        "source_paths": evidence["source_paths"],
        "naming_marker_policy": evidence.get("naming_marker_policy", {}),
        "annotation_evidence_policy": evidence.get("annotation_evidence_policy", {}),
        "user_constraints": evidence.get("annotation_evidence_policy", {}).get("user_constraints", {}),
        "deterministic_tnk_arbitration": evidence.get("deterministic_tnk_arbitration", {}),
        "full_evidence_pack": "annotation_evidence_pack.json",
        "usage": "Annotate from this digest first; open the full evidence pack only for a targeted conflict.",
    }


def main():
    parser = argparse.ArgumentParser(description="Validate paths/metadata and create a compact annotation evidence pack.")
    parser.add_argument("--avg", required=True)
    parser.add_argument("--markers", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--workspace-root", required=True)
    parser.add_argument("--species", required=True)
    parser.add_argument("--tissue", required=True)
    parser.add_argument("--experimental-system", default="")
    parser.add_argument("--context-source", action="append", default=[])
    parser.add_argument("--annotation-level", choices=["major"], required=True)
    parser.add_argument("--parent-population", required=True)
    parser.add_argument("--parent-kind", choices=["auto", "lineage", "state", "mixed", "unknown"], default="auto")
    parser.add_argument("--ratios", help="Optional full gene-by-cluster detection-ratio table (TSV/CSV).")
    parser.add_argument("--gene-map", help="Optional source_gene-to-canonical_gene TSV for cross-species symbols.")
    parser.add_argument("--cell-evidence", help="Optional per-cluster cell-level validation JSON.")
    parser.add_argument("--umap", help="Optional UMAP/embedding image for topology review.")
    parser.add_argument("--evidence-config", help="Optional versioned deterministic evidence configuration JSON.")
    parser.add_argument("--knowledge-base", help="Optional approved shared annotation knowledge-base JSON.")
    parser.add_argument("--project-prior", help="Optional auditable XLSX project label prior with cluster and label columns.")
    parser.add_argument("--project-major-label", action="append", default=[], help="Allowed project major label; repeat as needed.")
    parser.add_argument("--context-json", help="Optional sample context JSON: age, sex, disease, treatment, anatomy, platform, depth, and doublet metadata.")
    parser.add_argument("--annotation-constraints", help="Optional JSON with exclude_labels/conflict_markers and per-cluster constraints.")
    parser.add_argument("--exclude-label", action="append", default=[], help="Hard-exclude a final major identity label; repeat as needed.")
    parser.add_argument("--exclude-marker", action="append", default=[], help="Treat a marker as conflict/contamination evidence, not positive identity/state evidence; repeat as needed.")
    parser.add_argument("--allow-partial-ratios", action="store_true", help="Allow a marker-only ratio table for provisional review; formal full-ratio claims remain disabled.")
    parser.add_argument("--top-n", type=int, default=60)
    parser.add_argument("--informative-n", type=int, default=25)
    args = parser.parse_args()

    workspace = assert_e_drive(args.workspace_root, "workspace root")
    avg = assert_e_drive(args.avg, "average-expression input")
    markers = assert_e_drive(args.markers, "marker input")
    ratios = assert_e_drive(args.ratios, "expression-ratio input") if args.ratios else None
    gene_map = assert_e_drive(args.gene_map, "gene-map input") if args.gene_map else None
    cell_evidence = assert_e_drive(args.cell_evidence, "cell-evidence input") if args.cell_evidence else None
    umap = assert_e_drive(args.umap, "UMAP input") if args.umap else None
    evidence_config = assert_e_drive(args.evidence_config, "evidence config") if args.evidence_config else None
    knowledge_base = assert_e_drive(args.knowledge_base, "knowledge base") if args.knowledge_base else None
    project_prior_path = assert_e_drive(args.project_prior, "project prior") if args.project_prior else None
    context_path = assert_e_drive(args.context_json, "sample context") if args.context_json else None
    constraints_path = assert_e_drive(args.annotation_constraints, "annotation constraints") if args.annotation_constraints else None
    output_dir = assert_within(assert_e_drive(args.output_dir, "output directory"), workspace, "output directory")
    if not avg.is_file() or not markers.is_file():
        raise FileNotFoundError(f"Input file missing: avg={avg.exists()}, markers={markers.exists()}")
    for role, optional in (("ratios", ratios), ("gene map", gene_map), ("cell evidence", cell_evidence), ("UMAP", umap), ("evidence config", evidence_config), ("knowledge base", knowledge_base), ("project prior", project_prior_path), ("sample context", context_path), ("annotation constraints", constraints_path)):
        if optional is not None and not optional.is_file():
            raise FileNotFoundError(f"Optional {role} input missing: {optional}")
    reparse = has_reparse_component(output_dir.parent, workspace)
    if reparse:
        raise ValueError(f"Output path contains a symlink/junction/reparse point, which can break sandbox refresh: {reparse}")
    output_dir.mkdir(parents=True, exist_ok=True)

    parent_kind = infer_parent_kind(args.parent_population, args.parent_kind)
    project_prior = load_project_prior(project_prior_path)
    sample_context = load_sample_context(context_path)
    annotation_constraints = load_annotation_constraints(
        constraints_path, args.exclude_label, args.exclude_marker
    )
    project_major_vocabulary = sorted({
        *(str(item).strip() for item in args.project_major_label if str(item).strip()),
        *(label for label in project_prior.values() if label),
    })
    evidence = build_evidence(avg, markers, args.top_n, args.informative_n, args.species)
    if evidence["missing_marker_clusters"] or evidence["extra_marker_clusters"]:
        raise ValueError("Average-expression clusters and marker-table clusters do not match")
    unknown_prior_clusters = sorted(set(project_prior) - {str(item) for item in evidence["clusters"]})
    if unknown_prior_clusters:
        raise ValueError(f"Project prior contains clusters absent from the expression inputs: {unknown_prior_clusters}")
    evidence = enrich_evidence(
        evidence,
        ratio_path=ratios,
        gene_map_path=gene_map,
        cell_evidence_path=cell_evidence,
        config_path=evidence_config,
        annotation_level=args.annotation_level,
        species=args.species,
        tissue=args.tissue,
        parent_population=args.parent_population,
        parent_kind=parent_kind,
        knowledge_base_path=knowledge_base,
        project_major_vocabulary=project_major_vocabulary,
        project_label_prior=project_prior,
        require_complete_ratio=bool(ratios and not args.allow_partial_ratios),
        sample_context=sample_context,
        user_constraints=annotation_constraints,
    )
    metadata = {
        "species": args.species, "tissue": args.tissue,
        "experimental_system": args.experimental_system,
        "context_sources": [str(assert_e_drive(path, "context source")) for path in args.context_source],
        "annotation_level": args.annotation_level,
        "parent_population": args.parent_population, "parent_kind": parent_kind,
        "project_major_vocabulary": project_major_vocabulary,
        "project_prior_clusters": sorted(project_prior),
        "sample_context": sample_context,
        "annotation_constraints": annotation_constraints,
        "ratio_mode": "full_ratio" if ratios and not args.allow_partial_ratios else ("partial_ratio" if ratios else "positive_markers_only"),
        "interpretation_rule": (
            "Parent population is a state-based bucket; resolve every coherent lineage and keep state separate from identity."
            if parent_kind == "state" else
            "Parent population is a true all-cell or mixed-lineage collection; resolve broad lineages independently."
            if parent_kind == "mixed" else
            "Parent population is lineage-constrained; retain the lineage and require coherent positive DE-marker evidence for off-lineage calls."
        ),
    }
    evidence["confirmed_metadata"] = metadata
    evidence["source_paths"] = {
        "cell_avg_exp": str(avg), "marker_table": str(markers),
        "expression_ratio_table": str(ratios) if ratios else "",
        "gene_map": str(gene_map) if gene_map else "",
        "cell_evidence": str(cell_evidence) if cell_evidence else "",
        "umap": str(umap) if umap else "",
        "knowledge_base": str(knowledge_base) if knowledge_base else evidence.get("annotation_evidence_policy", {}).get("knowledge_base_source", ""),
        "project_prior": str(project_prior_path) if project_prior_path else "",
        "context_json": str(context_path) if context_path else "",
        "annotation_constraints": str(constraints_path) if constraints_path else "",
    }
    evidence_path = output_dir / "annotation_evidence_pack.json"
    evidence_path.write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    digest_path = output_dir / "annotation_evidence_digest.json"
    digest_path.write_text(json.dumps(evidence_digest(evidence), ensure_ascii=False, indent=2), encoding="utf-8")
    template_path = output_dir / "annotation_records.template.json"
    template_path.write_text(json.dumps([template_record(c) for c in evidence["clusters"]], ensure_ascii=False, indent=2), encoding="utf-8")
    manifest = {
        "status": "prepared", "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "workspace_root": str(workspace), "output_dir": str(output_dir),
        "metadata": metadata, "average_reader": evidence["average_reader"],
        "average_matrix_semantics": evidence.get("average_matrix_semantics", {}),
        "input_shape": evidence["average_shape"], "cluster_count": len(evidence["clusters"]),
        "evidence_pack": str(evidence_path), "evidence_digest": str(digest_path),
        "annotation_template": str(template_path),
        "path_policy": {"write_drive": "E:", "c_drive_input_allowed": False, "junctions_in_output_allowed": False},
        "naming_marker_policy": evidence.get("naming_marker_policy", {}),
        "annotation_evidence_policy": evidence.get("annotation_evidence_policy", {}),
        "deterministic_tnk_arbitration": evidence.get("deterministic_tnk_arbitration", {}),
        "python": sys.executable,
    }
    manifest_path = output_dir / "annotation_run_manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "status": "prepared", "parent_kind": parent_kind, "clusters": evidence["clusters"],
        "average_reader": evidence["average_reader"],
        "average_gene_header": evidence.get("average_gene_header", "GeneName"),
        "average_matrix_semantics": evidence.get("average_matrix_semantics", {}),
        "evidence_mode": sorted({item["evidence_mode"] for item in evidence["deterministic_annotation_evidence"].values()}),
        "tnk_regime": evidence.get("deterministic_tnk_arbitration", {}).get("recommended_regime"),
        "evidence_bytes": evidence_path.stat().st_size, "digest_bytes": digest_path.stat().st_size,
        "evidence_pack": str(evidence_path), "evidence_digest": str(digest_path),
        "annotation_template": str(template_path),
        "manifest": str(manifest_path),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()



