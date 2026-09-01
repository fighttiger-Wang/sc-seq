#!/usr/bin/env python3
"""Build and structurally verify the standardized annotation workbook."""

import argparse
import json
import os
import platform
import re
import subprocess
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from annotation_override_policy import validate_identity_override


SKILL_NAME = "sc-major-celltype-annotation-auto"
SKILL_VERSION = "0.3.2"


def register_shared_case(workspace, records_path, evidence_path, qa_path, output):
    if re.search(r"(^|[_-])(test|regression|validation|builder|preflight)([_-]|$)", output.parent.name, re.I) or output.parent.name.lower().startswith("skill_"):
        return {"status": "skipped_nonproduction", "project_dir": str(output.parent)}
    script = workspace / "local-marketplace" / "shared" / "sc-annotation-case-registry" / "case_registry.py"
    sidecar = output.with_suffix(".case-registry.json")
    if not script.is_file():
        result = {"status": "failed", "error": f"Shared case registry script missing: {script}"}
    else:
        completed = subprocess.run(
            [sys.executable, str(script), "register", "--records", str(records_path), "--evidence", str(evidence_path),
             "--qa", str(qa_path), "--project-dir", str(output.parent), "--skill-name", SKILL_NAME,
             "--skill-version", SKILL_VERSION], text=True, capture_output=True
        )
        if completed.returncode == 0:
            try:
                result = json.loads(completed.stdout)
            except json.JSONDecodeError:
                result = {"status": "failed", "error": "Registry returned invalid JSON", "stdout": completed.stdout[-2000:]}
        else:
            result = {"status": "failed", "error": completed.stderr.strip() or completed.stdout.strip(), "returncode": completed.returncode}
    sidecar.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return result


FIELDS = [
    "cluster_id", "celltype_cn", "celltype_en", "broad_type", "fine_type", "state",
    "supporting_markers", "conflicting_markers", "candidate_labels", "confidence",
    "quality_score", "mixed_or_doublet", "mixture_type", "possible_components",
    "rationale", "manual_review", "review_action",
    "evidence_mode", "evidence_completeness", "primary_evidence_label", "primary_major_label",
    "primary_evidence_score", "runner_up_evidence_label", "runner_up_major_label",
    "runner_up_evidence_score", "score_margin", "positive_marker_coverage",
    "detection_specificity", "rival_lineage", "rival_major_label", "rival_lineage_score",
    "tnk_provisional", "risk_level", "recommended_action", "decision_trace",
]
HEADERS = [
    "Cluster", "建议中文名称", "English label", "大类谱系", "精细类型", "细胞状态",
    "关键 Marker", "冲突 Marker", "候选标签", "置信度", "质量评分", "疑似多细胞/双细胞",
    "混合类型", "可能组成", "判定依据", "人工复核", "人工复核建议",
]
HEADERS.extend([
    "Evidence_Mode", "Evidence_Completeness", "Primary_Evidence", "Primary_Major_Label",
    "Primary_Score", "Runner_Up_Evidence", "Runner_Up_Major_Label", "Runner_Up_Score",
    "Score_Margin", "Core_Marker_Coverage", "Detection_Specificity", "Rival_Lineage",
    "Rival_Major_Label", "Rival_Score", "T_NK_Provisional", "Risk_Level",
    "Recommended_Action", "Decision_Trace",
])
STRUCTURED_FIELDS = [
    "stable_id", "parent_path", "tissue_module", "developmental_stage", "ontology_node_kind",
    "tissue_scope", "tissue_scope_match", "tissue_context_review", "disease_role", "state_list", "primary_state",
    "cross_species_inference", "panel_species", "marker_panel_evidence_ids", "display_label",
    "mixed_evidence", "review_in_subcluster", "mixed_population", "suspected_doublet", "auto_merge_allowed",
    "user_constraint_audit",
]
STRUCTURED_HEADERS = [
    "Stable_ID", "Parent_Path", "Tissue_Module", "Developmental_Stage", "Ontology_Node_Kind",
    "Tissue_Scope", "Tissue_Scope_Match", "Tissue_Context_Review", "Disease_Role", "State_List", "Primary_State",
    "Cross_Species_Inference", "Panel_Species", "Marker_Panel_Evidence", "Display_Label",
    "Mixed_Evidence", "Review_In_Subcluster", "Mixed_Population", "Suspected_Doublet", "Auto_Merge_Allowed",
    "User_Constraint_Audit",
]
FIELDS[6:6] = STRUCTURED_FIELDS
HEADERS[6:6] = STRUCTURED_HEADERS
AUDIT_FIELDS = [
    "label_basis", "canonical_subtype", "top_marker_gene", "literature_source",
    "naming_grammar", "contextually_excluded_naming_markers",
]
LABEL_BASIS = {
    "canonical_subtype", "validated_external_candidate", "multi_cell_annotation",
    "top_marker_fallback", "feature_gene_fallback",
}
CONFIDENCE = {"high", "medium-high", "medium", "low"}
FINAL_LABEL_INVALID = re.compile(r"[^A-Za-z0-9_\u3400-\u4DBF\u4E00-\u9FFF]+")
T_NK_T_MARKERS = {"CD2", "CD3D", "CD3E", "CD3G", "CD5", "CD7", "CD8A", "CD8B", "IL7R", "TRAC", "TRAT1", "TRBC1", "TRBC2", "TRDC", "TRGC1", "TRGC2", "LCK", "LAT"}
T_NK_T_COMPLEX = {"CD3D", "CD3E", "CD3G", "TRAC", "TRBC1", "TRBC2", "TRDC", "TRGC1", "TRGC2"}
T_NK_NK_MARKERS = {"NKG7", "CCL5", "PRF1", "GZMB", "GZMH", "GNLY", "KLRD1", "FCGR3A", "NCAM1", "KLRF1", "KLRC1", "KLRC2", "NCR1", "NCR3"}
T_NK_NK_SPECIFIC = {"GNLY", "KLRD1", "FCGR3A", "NCAM1", "KLRF1", "KLRC1", "KLRC2", "NCR1", "NCR3"}

MAIN_FIELDS = [
    "cluster_id", "celltype_cn", "celltype_en", "broad_type", "developmental_stage", "state",
    "disease_role", "supporting_markers", "conflicting_markers", "candidate_labels", "confidence",
    "quality_score", "mixed_or_doublet", "mixture_type", "possible_components", "rationale",
    "manual_review", "review_action",
]
MAIN_HEADERS = [
    "Cluster", "中文名称", "Celltype_EN", "大类谱系", "发育阶段", "细胞状态",
    "疾病或组织角色", "关键 Marker", "冲突 Marker", "候选标签", "置信度",
    "质量评分", "疑似多细胞/双细胞", "混合类型", "可能组成", "判定依据",
    "需要复核", "复核建议",
]
EVIDENCE_FIELDS = [
    "cluster_id", "stable_id", "fine_type", "parent_path", "tissue_module", "ontology_node_kind",
    "tissue_scope", "tissue_scope_match", "tissue_context_review", "state_list", "primary_state",
    "cross_species_inference", "panel_species", "marker_panel_evidence_ids", "evidence_mode",
    "evidence_completeness", "primary_evidence_label", "primary_major_label", "primary_evidence_score",
    "runner_up_evidence_label", "runner_up_major_label", "runner_up_evidence_score", "score_margin",
    "positive_marker_coverage", "detection_specificity", "rival_lineage", "rival_major_label",
    "rival_lineage_score", "tnk_provisional", "risk_level", "recommended_action", "mixed_evidence",
    "review_in_subcluster", "mixed_population", "suspected_doublet", "auto_merge_allowed", "label_basis", "canonical_subtype", "literature_source",
    "naming_grammar",
    "literature_details", "override_validation", "override_audit", "user_constraint_audit",
]
EVIDENCE_HEADERS = [
    "Cluster", "Stable_ID", "精细类型", "本体路径", "组织模块", "本体节点类型",
    "适用组织", "组织匹配", "组织背景复核", "完整状态", "主要状态", "跨物种推断",
    "参考物种", "Marker 证据", "证据模式", "证据完整度", "主候选", "主候选大类",
    "主候选评分", "次候选", "次候选大类", "次候选评分", "评分差", "核心 Marker 覆盖",
    "检测特异性", "竞争谱系", "竞争大类", "竞争评分", "T/NK 判定", "风险等级",
    "处置建议", "混合证据", "亚群阶段复核", "混合群", "疑似双细胞", "允许自动合并", "命名依据", "标准亚型",
    "文献依据", "命名规则",
    "Structured literature details", "Override validation", "Override audit", "User Constraint Audit",
]


def inject_deterministic_evidence(records, evidence):
    decisions = evidence.get("deterministic_annotation_evidence", {})
    for record in records:
        cluster = str(record.get("cluster_id", ""))
        if cluster not in decisions:
            raise ValueError(f"Cluster {cluster} lacks deterministic annotation evidence")
        decision = decisions[cluster]
        external_candidate = record.get("label_basis") == "validated_external_candidate"
        for field in (
            "evidence_mode", "evidence_completeness", "primary_evidence_label", "primary_major_label",
            "primary_evidence_score", "runner_up_evidence_label", "runner_up_major_label",
            "runner_up_evidence_score", "score_margin", "positive_marker_coverage",
            "detection_specificity", "rival_lineage", "rival_major_label", "rival_lineage_score",
            "tnk_provisional", "risk_level", "recommended_action",
        ):
            record[field] = decision[field]
        for field in (
            "stable_id", "developmental_stage", "ontology_node_kind", "tissue_scope_match",
            "tissue_context_review", "primary_state", "cross_species_inference", "panel_species",
            "display_label", "mixed_population", "suspected_doublet", "auto_merge_allowed",
            "mixed_evidence", "review_in_subcluster",
            "user_constraint_audit",
        ):
            preserve_explicit_false = (
                external_candidate
                and field in {"cross_species_inference", "auto_merge_allowed"}
                and field in record
                and isinstance(record[field], bool)
            )
            if (not external_candidate or not record.get(field)) and not preserve_explicit_false:
                record[field] = decision.get(field, "")
        record["user_constraint_audit"] = json.dumps(
            decision.get("user_constraint_audit", {}), ensure_ascii=False, sort_keys=True
        )
        for field in ("parent_path", "tissue_module", "tissue_scope", "disease_role", "state_list", "marker_panel_evidence_ids"):
            if not external_candidate or not record.get(field):
                record[field] = json.dumps(decision.get(field, []), ensure_ascii=False)
        if external_candidate:
            record["display_label"] = (
                record.get("display_label")
                or record.get("celltype_en")
                or record.get("stable_id", "")
            )
            record["manual_review"] = True
        if not record.get("state"):
            record["state"] = decision.get("primary_state", "")
        if decision.get("mixed_population"):
            record["mixed_or_doublet"] = True
            record["manual_review"] = True
            record["stable_id"] = "Multi_cell"
            record["display_label"] = "Multi_cell"
            record["celltype_en"] = "Multi_cell"
            record["celltype_cn"] = "多细胞"
            record["broad_type"] = "Multi_cell"
            record["fine_type"] = "Multi_cell"
            record["canonical_subtype"] = "Multi_cell"
            record["label_basis"] = "multi_cell_annotation"
            record["mixture_type"] = decision.get("mixture_type") or record.get("mixture_type") or "mixed_population/suspected_doublet"
            components = decision.get("possible_components", [])
            if components:
                record["possible_components"] = "; ".join(str(item) for item in components)
        record["decision_trace"] = json.dumps(decision.get("decision_trace", {}), ensure_ascii=False, sort_keys=True)


def cluster_sort_key(value):
    """Sort numeric cluster IDs by value, then nonnumeric IDs naturally."""
    text = str(value).strip()
    try:
        number = Decimal(text)
        if number.is_finite():
            return (0, number, text)
    except InvalidOperation:
        pass
    natural = tuple((0, int(part)) if part.isdigit() else (1, part.lower()) for part in re.split(r"(\d+)", text))
    return (1, natural, text)


def resolved_e(path, role):
    resolved = Path(path).resolve()
    if platform.system() == "Windows" and os.path.splitdrive(str(resolved))[0].upper() != "E:":
        raise ValueError(f"{role} must be on E: {resolved}")
    return resolved


def normalize_final_label(value):
    """Replace spaces and nonportable label characters with one underscore."""
    text = FINAL_LABEL_INVALID.sub("_", str(value).strip())
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        raise ValueError(f"Final label becomes empty after normalization: {value!r}")
    return text


def normalize_record_labels(records):
    changes = []
    for record in records:
        for field in ("celltype_cn", "celltype_en"):
            original = str(record[field])
            normalized = normalize_final_label(original)
            record[field] = normalized
            if normalized != original:
                changes.append({
                    "cluster_id": str(record["cluster_id"]),
                    "field": field,
                    "original": original,
                    "normalized": normalized,
                })
    return changes

def within(path, root, role):
    path, root = Path(path).resolve(), Path(root).resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"{role} must be inside workspace root {root}: {path}") from exc
    return path


def validate(records, clusters, evidence):
    errors = []
    ratio_validation = evidence.get("annotation_evidence_policy", {}).get("ratio_validation", {})
    if ratio_validation.get("provided") and not ratio_validation.get("complete"):
        errors.append(
            "Formal major annotation requires a complete full-ratio table; "
            "partial-ratio evidence is provisional only"
        )
    record_clusters = [str(r.get("cluster_id", "")) for r in records]
    expected_clusters = sorted((str(c) for c in clusters), key=cluster_sort_key)
    if record_clusters != expected_clusters:
        errors.append(f"Cluster order/membership mismatch: expected={expected_clusters}, records={record_clusters}")
    top_by_cluster = {}
    informative_by_cluster = {}
    for cluster in clusters:
        profile = evidence["cluster_profiles"][str(cluster)]
        naming_top = profile.get("naming_top_marker")
        top_by_cluster[str(cluster)] = str(naming_top["gene"]) if naming_top else None
        informative_by_cluster[str(cluster)] = [
            str(marker["gene"]) for marker in profile.get("top_informative_markers", [])
        ]
    grammars = set()
    canonical_counts = {}
    for index, record in enumerate(records):
        missing = [field for field in FIELDS + AUDIT_FIELDS if field not in record]
        if missing:
            errors.append(f"Record {index} missing fields: {missing}")
            continue
        cluster_id = str(record["cluster_id"])
        deterministic_decision = evidence.get("deterministic_annotation_evidence", {}).get(cluster_id, {})
        constraint_audit = deterministic_decision.get("user_constraint_audit", {})
        if constraint_audit.get("final_identity_excluded"):
            errors.append(
                f"Cluster {cluster_id} final identity violates an explicit user exclusion: "
                f"{constraint_audit.get('selected_final_identity', '')}"
            )
        try:
            recorded_constraints = json.loads(str(record.get("user_constraint_audit", "{}")))
        except json.JSONDecodeError:
            recorded_constraints = {}
        excluded_labels = {str(item).strip() for item in recorded_constraints.get("exclude_labels", []) if str(item).strip()}
        if str(record.get("stable_id", "")).strip() in excluded_labels or str(record.get("celltype_en", "")).strip() in excluded_labels:
            errors.append(f"Cluster {cluster_id} final record uses an explicitly excluded identity")
        override_audit = validate_identity_override(record, deterministic_decision)
        record["override_audit"] = override_audit
        errors.extend(override_audit["errors"])
        expected_top = top_by_cluster[cluster_id]
        selected_marker = str(record["top_marker_gene"])
        contextual_exclusions = record["contextually_excluded_naming_markers"]
        if record["label_basis"] not in LABEL_BASIS:
            errors.append(f"Cluster {cluster_id} invalid label_basis: {record['label_basis']}")
        elif record["label_basis"] not in {"canonical_subtype", "validated_external_candidate", "multi_cell_annotation"}:
            errors.append(
                f"Cluster {cluster_id} major annotation requires a canonical subtype "
                "or a validated external candidate"
            )
        if not isinstance(contextual_exclusions, list):
            errors.append(f"Cluster {cluster_id} contextually_excluded_naming_markers must be a list")
            contextual_exclusions = []
        if expected_top is None:
            if record["label_basis"] in {"top_marker_fallback", "feature_gene_fallback"}:
                errors.append(
                    f"Cluster {cluster_id} has no standardized informative marker; "
                    "marker fallback is not allowed and requires a canonical label or manual review"
                )
        elif record["label_basis"] != "feature_gene_fallback" and selected_marker != expected_top:
            errors.append(
                f"Cluster {cluster_id} top_marker_gene must equal highest-log2FC standardized "
                f"informative naming marker {expected_top}"
            )
        if record["label_basis"] == "feature_gene_fallback" and expected_top is not None:
            informative = informative_by_cluster[cluster_id]
            if selected_marker not in informative:
                errors.append(
                    f"Cluster {cluster_id} feature fallback marker {selected_marker} must occur in top_informative_markers"
                )
            else:
                preceding = informative[:informative.index(selected_marker)]
                excluded_genes = []
                for item in contextual_exclusions:
                    if not isinstance(item, dict) or not str(item.get("gene", "")).strip() or not str(item.get("reason", "")).strip():
                        errors.append(f"Cluster {cluster_id} contextual exclusions require non-empty gene and reason")
                        continue
                    excluded_genes.append(str(item["gene"]))
                if excluded_genes != preceding:
                    errors.append(
                        f"Cluster {cluster_id} contextual exclusions must exactly match preceding informative markers: "
                        f"expected={preceding}, recorded={excluded_genes}"
                    )
            if not record["manual_review"]:
                errors.append(f"Cluster {cluster_id} feature_gene_fallback requires manual_review=true")
            if record["confidence"] == "high":
                errors.append(f"Cluster {cluster_id} feature_gene_fallback confidence cannot be high")
        elif contextual_exclusions:
            errors.append(f"Cluster {cluster_id} contextual exclusions are allowed only for feature_gene_fallback")
        if not str(record["naming_grammar"]).strip():
            errors.append(f"Cluster {cluster_id} lacks naming_grammar")
        else:
            grammars.add(str(record["naming_grammar"]).strip())
        canonical = str(record["canonical_subtype"]).strip()
        stable_id = str(record.get("stable_id", "")).strip()
        if deterministic_decision.get("resolution_search_required"):
            errors.append(
                f"Cluster {cluster_id} remains unresolved and requires targeted research; formal delivery is blocked"
            )
        if not stable_id:
            errors.append(f"Cluster {cluster_id} lacks a defensible final Stable_ID")
        if stable_id == "Cell" or str(record.get("celltype_en", "")).strip() == "Cell":
            errors.append(f"Cluster {cluster_id} cannot use forbidden final label Cell")
        if record["label_basis"] == "multi_cell_annotation":
            if stable_id != "Multi_cell" or str(record.get("celltype_en", "")).strip() != "Multi_cell":
                errors.append(f"Cluster {cluster_id} multi-cell annotation must use Stable_ID and English label Multi_cell")
            if str(record.get("celltype_cn", "")).strip() != "多细胞":
                errors.append(f"Cluster {cluster_id} Multi_cell annotation must use Chinese name 多细胞")
            if not record.get("mixed_population") or record.get("auto_merge_allowed"):
                errors.append(f"Cluster {cluster_id} Multi_cell annotation requires mixed_population=true and auto_merge_allowed=false")
            if not str(record.get("possible_components", "")).strip():
                errors.append(f"Cluster {cluster_id} Multi_cell annotation requires possible_components")
            if canonical != "Multi_cell":
                errors.append(f"Cluster {cluster_id} Multi_cell annotation requires canonical_subtype=Multi_cell")
        elif record["label_basis"] == "canonical_subtype":
            if not canonical:
                errors.append(f"Cluster {cluster_id} canonical_subtype label lacks canonical_subtype")
            if not str(record["literature_source"]).strip():
                errors.append(f"Cluster {cluster_id} canonical subtype lacks literature_source")
            canonical_counts[canonical] = canonical_counts.get(canonical, 0) + 1
        elif record["label_basis"] == "validated_external_candidate":
            stable_id = str(record.get("stable_id", "")).strip()
            sources = [item.strip() for item in re.split(r"[;；]", str(record["literature_source"])) if item.strip()]
            if not canonical or canonical != stable_id:
                errors.append(f"Cluster {cluster_id} validated external candidate requires canonical_subtype=stable_id")
            if len(sources) < 2:
                errors.append(f"Cluster {cluster_id} validated external candidate requires at least two independent sources")
            if not record["manual_review"]:
                errors.append(f"Cluster {cluster_id} validated external candidate requires manual_review=true")
            if record["confidence"] == "high":
                errors.append(f"Cluster {cluster_id} validated external candidate cannot receive high confidence before regression promotion")
        elif record["label_basis"] in {"top_marker_fallback", "feature_gene_fallback"}:
            if canonical:
                errors.append(f"Cluster {cluster_id} marker fallback must leave canonical_subtype blank")
            if str(record["literature_source"]).strip():
                errors.append(f"Cluster {cluster_id} marker fallback must leave literature_source blank")
            if expected_top is not None and not str(record["celltype_en"]).startswith(selected_marker + "_"):
                errors.append(f"Cluster {cluster_id} fallback label must start with {selected_marker}_")
        if record["confidence"] not in CONFIDENCE:
            errors.append(f"Cluster {record['cluster_id']} invalid confidence: {record['confidence']}")
        score = record["quality_score"]
        if not isinstance(score, (int, float)) or not 0 <= score <= 100:
            errors.append(f"Cluster {record['cluster_id']} quality_score must be 0-100")
        if not record["celltype_cn"] or not record["celltype_en"] or not record["rationale"]:
            errors.append(f"Cluster {record['cluster_id']} lacks a final label or rationale")
        if record["risk_level"] != "R0_ACCEPT" and not record["manual_review"]:
            errors.append(f"Cluster {cluster_id} deterministic risk {record['risk_level']} requires manual_review=true")
        if record.get("mixed_population") and (record.get("auto_merge_allowed") or not record.get("mixed_or_doublet")):
            errors.append(f"Cluster {cluster_id} mixed population must block automatic merging and carry mixed_or_doublet=true")
        if record.get("mixed_population") and stable_id != "Multi_cell":
            errors.append(f"Cluster {cluster_id} mixed population must use final label Multi_cell")
        if record["evidence_mode"] == "minimal" and record["confidence"] == "high":
            errors.append(f"Cluster {cluster_id} positive-marker-only evidence cannot receive high confidence")
        if "JCHAIN" in str(record["supporting_markers"]).upper() and "plasma" in str(record["celltype_en"]).lower():
            coherent = sum(g in str(record["supporting_markers"]).upper() for g in ["PRDM1", "XBP1", "SDC1", "MZB1", "DERL3"])
            if coherent < 3:
                errors.append(f"Cluster {record['cluster_id']} plasma label lacks coherent secretion program")
        if canonical == "T_NK_cell" or str(record["celltype_en"]).strip() == "T_NK_cell":
            if str(record["celltype_en"]).strip() != "T_NK_cell" or str(record["broad_type"]).strip() != "T_NK_cell" or canonical != "T_NK_cell":
                errors.append(f"Cluster {cluster_id} composite label must match celltype_en, broad_type, and canonical_subtype")
    deterministic_tnk = evidence.get("deterministic_tnk_arbitration", {})
    expected_regime = deterministic_tnk.get("recommended_regime")
    if expected_regime != "per_cluster":
        errors.append(f"Unsupported legacy T/NK arbitration regime: {expected_regime}")
    provisional = deterministic_tnk.get("provisional_by_cluster", {})
    for record in records:
        cluster_id = str(record["cluster_id"])
        origin = provisional.get(cluster_id, "not_T_NK")
        final_label = str(record["celltype_en"]).strip()
        if origin == "T_supported" and not record.get("mixed_population") and final_label != "T_cell":
            errors.append(f"Cluster {cluster_id} deterministic T-supported evidence requires T_cell")
        if origin == "NK_supported" and not record.get("mixed_population") and final_label != "NK_cell":
            errors.append(f"Cluster {cluster_id} deterministic NK-supported evidence requires NK_cell")
        if origin == "unresolved_T_NK":
            if not (record.get("mixed_population") or record.get("mixed_evidence")) or record.get("auto_merge_allowed"):
                errors.append(f"Cluster {cluster_id} unresolved T/NK must carry mixed evidence and be blocked from automatic merging")
            if not record.get("manual_review"):
                errors.append(f"Cluster {cluster_id} unresolved T/NK requires manual_review=true")
    if len(grammars) > 1:
        errors.append(f"Multiple naming grammars in one table: {sorted(grammars)}")
    # Major-celltype mappings intentionally allow repeated canonical broad labels.
    if errors:
        raise ValueError("\n".join(errors))


def set_widths(ws, widths):
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def human_value(value):
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (list, tuple, set)):
        return "、".join(human_value(item) for item in value if human_value(item))
    if isinstance(value, dict):
        return "；".join(f"{key}：{human_value(item)}" for key, item in value.items() if human_value(item))
    text = str(value).strip()
    if text.startswith(("[", "{")):
        try:
            return human_value(json.loads(text))
        except json.JSONDecodeError:
            pass
    text = text.replace("[", "").replace("]", "").replace("{", "").replace("}", "")
    return re.sub(r"\s+", " ", text)


def compact_rows(ws, min_row, max_col, maximum_height=54):
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=max_col):
        longest = max((len(str(cell.value or "")) for cell in row), default=0)
        ws.row_dimensions[row[0].row].height = min(maximum_height, 22 + 11 * min(3, longest // 45))
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_header(ws, cell_range):
    fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--records", required=True)
    parser.add_argument("--evidence", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--workspace-root", required=True)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    workspace = resolved_e(args.workspace_root, "workspace root")
    records_path = within(resolved_e(args.records, "records"), workspace, "records")
    evidence_path = within(resolved_e(args.evidence, "evidence"), workspace, "evidence")
    output = within(resolved_e(args.output, "workbook output"), workspace, "workbook output")
    if output.exists() and not args.force:
        raise FileExistsError(f"Refusing to overwrite existing workbook without --force: {output}")
    records = json.loads(records_path.read_text(encoding="utf-8"))
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    records = sorted(records, key=lambda record: cluster_sort_key(record.get("cluster_id", "")))
    inject_deterministic_evidence(records, evidence)
    sorted_clusters = sorted((str(cluster) for cluster in evidence["clusters"]), key=cluster_sort_key)
    validate(records, sorted_clusters, evidence)
    label_normalization_changes = normalize_record_labels(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "注释结果"
    detail = wb.create_sheet("详细证据")
    source = wb.create_sheet("说明与数据来源")
    research = wb.create_sheet("细胞类型与文献")
    ws.append(MAIN_HEADERS)
    for record in records:
        ws.append([human_value(record.get(field, "")) for field in MAIN_FIELDS])
    detail.append(EVIDENCE_HEADERS)
    for record in records:
        detail.append([human_value(record.get(field, "")) for field in EVIDENCE_FIELDS])
    ws.freeze_panes = "A2"
    detail.freeze_panes = "A2"
    last_column = get_column_letter(len(MAIN_FIELDS))
    detail_last_column = get_column_letter(len(EVIDENCE_FIELDS))
    ws.sheet_view.showGridLines = False
    detail.sheet_view.showGridLines = False
    style_header(ws, f"A1:{last_column}1")
    style_header(detail, f"A1:{detail_last_column}1")
    ws.row_dimensions[1].height = 30
    detail.row_dimensions[1].height = 30
    thin = Side(style="thin", color="D9E2F3")
    for table, field_count in ((ws, len(MAIN_FIELDS)), (detail, len(EVIDENCE_FIELDS))):
        compact_rows(table, 2, field_count)
        for row in table.iter_rows(min_row=2, max_row=table.max_row, min_col=1, max_col=field_count):
            for cell in row:
                cell.border = Border(bottom=thin)
    set_widths(ws, [10, 22, 20, 16, 15, 15, 18, 28, 24, 24, 11, 11, 16, 18, 22, 42, 12, 36])
    set_widths(detail, [10, 18, 18, 34, 20, 18, 24, 14, 16, 28, 16, 14, 14, 28, 16, 18, 20, 18, 14, 20, 18, 14, 12, 16, 16, 18, 18, 14, 16, 18, 34, 14, 14, 16, 20, 20, 34, 24])
    quality_col = get_column_letter(MAIN_FIELDS.index("quality_score") + 1)
    mixed_col = get_column_letter(MAIN_FIELDS.index("mixed_or_doublet") + 1)
    review_col = get_column_letter(MAIN_FIELDS.index("manual_review") + 1)
    confidence_col = get_column_letter(MAIN_FIELDS.index("confidence") + 1)
    ws.conditional_formatting.add(f"{quality_col}2:{quality_col}{ws.max_row}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
    ws.conditional_formatting.add(f"{mixed_col}2:{mixed_col}{ws.max_row}", FormulaRule(formula=[f'{mixed_col}2="是"'], fill=PatternFill("solid", fgColor="F8CBAD")))
    ws.conditional_formatting.add(f"{review_col}2:{review_col}{ws.max_row}", FormulaRule(formula=[f'{review_col}2="是"'], fill=PatternFill("solid", fgColor="FFF2CC")))
    ws.conditional_formatting.add(f"{confidence_col}2:{confidence_col}{ws.max_row}", FormulaRule(formula=[f'{confidence_col}2="low"'], fill=PatternFill("solid", fgColor="FCE4D6")))
    multi_cell_red = PatternFill("solid", fgColor="FFF8696B")
    warning_fields = (
        "mixed_population", "mixed_or_doublet", "suspected_doublet", "debris",
        "low_quality", "background_interference", "abnormal_state", "abnormality_flags",
        "interpretation_flags", "plotting_warning", "remove_cluster", "needs_decontamination",
    )

    def has_annotation_warning(record):
        if record.get("stable_id") == "Multi_cell":
            return True
        for field in warning_fields:
            value = record.get(field)
            if isinstance(value, (list, tuple, set, dict)) and value:
                return True
            if isinstance(value, str) and value.strip().lower() not in {"", "no", "false", "否", "none", "nan"}:
                return True
            if isinstance(value, bool) and value:
                return True
        return False

    for row_index, record in enumerate(records, start=2):
        if has_annotation_warning(record):
            name_cell = ws.cell(row_index, MAIN_FIELDS.index("celltype_cn") + 1)
            name_cell.fill = multi_cell_red
            name_cell.font = Font(bold=True, color="FFFFFFFF")

    metadata = evidence.get("confirmed_metadata", {})
    paths = evidence.get("source_paths", {})
    source_rows = [
        ["项目", f"{metadata.get('species', '')} {metadata.get('tissue', '')} {metadata.get('parent_population', '')} 大类细胞注释"],
        ["物种", metadata.get("species", "")], ["组织/解剖来源", metadata.get("tissue", "")],
        ["注释层级", metadata.get("annotation_level", "")], ["父群", metadata.get("parent_population", "")],
        ["父群类型", metadata.get("parent_kind", "")], ["父群解释规则", metadata.get("interpretation_rule", "")],
        ["平均表达源文件", Path(paths.get("cell_avg_exp", "")).name],
        ["Marker源文件", Path(paths.get("marker_table", "")).name],
        ["输入规模", f"{evidence['average_shape'][0]:,} genes × {evidence['average_shape'][1]} clusters"],
        ["平均表达读取", evidence.get("average_reader", "")],
        ["命名模式", "使用通行短名称；身份、发育阶段和状态分别记录"],
        ["候选策略", "开放生成候选，使用组合 Marker、冲突证据和数据集内参照严格验证"],
        ["跨物种策略", "缺少物种专属面板时使用同源基因与保守表达程序推断并降低置信度"],
        ["重点复核", "、".join(str(r["cluster_id"]) for r in records if r["manual_review"]) or "无"],
    ]
    source_rows.extend([
        ["完整表达占比源文件", Path(paths.get("expression_ratio_table", "")).name or "未提供"],
        ["跨物种基因映射", Path(paths.get("gene_map", "")).name or "未提供"],
        ["单细胞验证证据", Path(paths.get("cell_evidence", "")).name or "未提供"],
        ["证据核心版本", f"core {evidence.get('annotation_evidence_policy', {}).get('core_version', '')}；config {evidence.get('annotation_evidence_policy', {}).get('config_version', '')}"],
    ])
    for row in source_rows:
        source.append(row)
    source.sheet_view.showGridLines = False
    set_widths(source, [24, 72])
    for row in source.iter_rows():
        row[0].fill = PatternFill("solid", fgColor="D9EAF7")
        row[0].font = Font(bold=True, color="17365D")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        source.row_dimensions[row[0].row].height = 28
    style_header(source, "A1:B1")

    research_headers = ["cluster", "细胞类型", "用于注释的特征基因", "候选/竞争类型", "文献来源", "证据来源ID", "UMAP判断"]
    research.append(research_headers)
    source_registry = evidence.get("annotation_evidence_policy", {}).get("evidence_source_registry", {})
    for record in records:
        source_text = record.get("literature_source", "")
        evidence_ids = record.get("marker_panel_evidence_ids", [])
        if not source_text and evidence_ids:
            source_text = "；".join(str(source_registry.get(item, {}).get("title", item)) for item in evidence_ids)
        research.append([
            record.get("cluster_id", ""),
            record.get("celltype_en", ""),
            record.get("supporting_markers", ""),
            record.get("candidate_labels", ""),
            source_text,
            "；".join(map(str, evidence_ids)),
            record.get("rationale", ""),
        ])
    research.sheet_view.showGridLines = False
    set_widths(research, [12, 24, 48, 36, 72, 28, 72])
    style_header(research, "A1:G1")
    for row in research.iter_rows(min_row=2, max_row=research.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        research.row_dimensions[row[0].row].height = 42

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    check = load_workbook(output, read_only=False, data_only=False)
    expected_sheets = ["注释结果", "详细证据", "说明与数据来源", "细胞类型与文献"]
    if check.sheetnames != expected_sheets:
        raise RuntimeError(f"Workbook sheet verification failed: {check.sheetnames}")
    if check["注释结果"].max_row != len(records) + 1 or check["详细证据"].max_row != len(records) + 1:
        raise RuntimeError("Workbook row-count verification failed")
    result_order = [str(check["注释结果"].cell(row, 1).value) for row in range(2, len(records) + 2)]
    detail_order = [str(check["详细证据"].cell(row, 1).value) for row in range(2, len(records) + 2)]
    if result_order != sorted_clusters or detail_order != sorted_clusters:
        raise RuntimeError(
            f"Workbook cluster-order verification failed: expected={sorted_clusters}, "
            f"result={result_order}, detail={detail_order}"
        )
    qa = {
        "status": "pass", "workbook": str(output), "sheets": check.sheetnames,
        "record_count": len(records), "review_clusters": [str(r["cluster_id"]) for r in records if r["manual_review"]],
        "research_sheet": "细胞类型与文献",
        "mixed_or_doublet_clusters": [str(r["cluster_id"]) for r in records if r["mixed_or_doublet"]],
        "multi_cell_clusters": [str(r["cluster_id"]) for r in records if str(r.get("stable_id")) == "Multi_cell"],
        "multi_cell_chinese_static_red": True,
        "t_nk_composite_clusters": [str(r["cluster_id"]) for r in records if str(r["celltype_en"]) == "T_NK_cell"],
        "t_nk_label_regime": ("composite" if any(str(r["celltype_en"]) == "T_NK_cell" for r in records) else ("split" if any(str(r["celltype_en"]) in {"T_cell", "NK_cell"} for r in records) else "not_applicable")),
        "formula_cells": 0, "source_files_unchanged": True,
        "visual_qa": "compact_human_record_layout",
        "auto_filter_enabled": False,
        "main_fields": MAIN_FIELDS,
        "evidence_fields": EVIDENCE_FIELDS,
        "cluster_order": sorted_clusters,
        "cluster_order_policy": "numeric_ascending_then_natural_alphanumeric",
        "naming_grammars": sorted({str(r["naming_grammar"]) for r in records}),
        "final_label_policy": "letters_digits_cjk_underscore_only",
        "final_labels_normalized": True,
        "label_normalization_changes": label_normalization_changes,
        "label_basis_counts": {
            "canonical_subtype": sum(r["label_basis"] == "canonical_subtype" for r in records),
            "validated_external_candidate": sum(r["label_basis"] == "validated_external_candidate" for r in records),
            "top_marker_fallback": sum(r["label_basis"] == "top_marker_fallback" for r in records),
            "feature_gene_fallback": sum(r["label_basis"] == "feature_gene_fallback" for r in records),
        },
        "naming_marker_policy": evidence.get("naming_marker_policy", {}),
        "naming_marker_by_cluster": {
            str(cluster): (
                evidence["cluster_profiles"][str(cluster)].get("naming_top_marker") or {}
            ).get("gene", "")
            for cluster in sorted_clusters
        },
        "selected_label_marker_by_cluster": {
            str(record["cluster_id"]): str(record["top_marker_gene"]) for record in records
        },
        "contextually_excluded_naming_markers_by_cluster": {
            str(record["cluster_id"]): record["contextually_excluded_naming_markers"] for record in records
        },
        "raw_top_marker_by_cluster": {
            str(cluster): (
                evidence["cluster_profiles"][str(cluster)].get("raw_top_marker") or {}
            ).get("gene", "")
            for cluster in sorted_clusters
        },
        "excluded_naming_markers_by_cluster": {
            str(cluster): evidence["cluster_profiles"][str(cluster)].get("excluded_naming_markers", [])
            for cluster in sorted_clusters
        },
        "annotation_evidence_policy": evidence.get("annotation_evidence_policy", {}),
        "deterministic_risk_by_cluster": {str(record["cluster_id"]): record["risk_level"] for record in records},
        "deterministic_action_by_cluster": {str(record["cluster_id"]): record["recommended_action"] for record in records},
        "override_audits_by_cluster": {
            str(record["cluster_id"]): record.get("override_audit", {}) for record in records
        },
        "deterministic_tnk_arbitration": evidence.get("deterministic_tnk_arbitration", {}),
    }
    qa_path = output.with_suffix(".qa.json")
    qa_path.write_text(json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8")
    qa["case_registry"] = register_shared_case(workspace, records_path, evidence_path, qa_path, output)
    print(json.dumps(qa, ensure_ascii=False))


if __name__ == "__main__":
    main()
