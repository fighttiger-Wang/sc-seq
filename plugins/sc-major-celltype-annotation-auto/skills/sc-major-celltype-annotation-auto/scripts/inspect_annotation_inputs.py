#!/usr/bin/env python3
"""Validate paired annotation workbooks and create a compact evidence pack."""

import argparse
import json
import math
import re
import statistics
import zipfile
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

SIGNATURES = {
    "B_lineage": ["Cd79a", "Cd79b", "Ebf1", "Pax5", "Ms4a1", "Cd74", "Fcmr", "Ighm", "Spib", "Fcrla"],
    "early_B": ["Rag1", "Rag2", "Dntt", "Igll1", "Vpreb1", "Vpreb2", "Vpreb3", "Il7r", "Ebf1", "Lef1"],
    "mature_B": ["Ms4a1", "Cd74", "Fcmr", "Tnfrsf13c", "Cd72", "Ighm", "Ighd", "Fcer2", "Cr2"],
    "plasma": ["Prdm1", "Xbp1", "Sdc1", "Mzb1", "Jchain", "Derl3", "Fkbp11", "Sec11c"],
    "T_lineage": ["Cd3d", "Cd3e", "Cd3g", "Trac", "Lck", "Lat", "Il7r"],
    "NK": ["Nkg7", "Klrd1", "Prf1", "Gzmb", "Gzmk", "Ccl5"],
    "myeloid": ["Lyz2", "Lyz", "Lst1", "Tyrobp", "Fcer1g", "Ctss", "Aif1", "Cst3"],
    "neutrophil": ["S100a8", "S100a9", "Retnlg", "Camp", "Ngp", "Pglyrp1", "Lcn2", "Csf3r"],
    "macrophage": ["C1qa", "C1qb", "C1qc", "Apoe", "Cd68", "Mertk", "Lpl"],
    "endothelial": ["Pecam1", "Vwf", "Kdr", "Emcn", "Eng", "Ramp2"],
    "stromal": ["Col1a1", "Col1a2", "Col3a1", "Dcn", "Lum", "Pdgfra", "Rgs5", "Acta2"],
    "erythroid": ["Hba-a1", "Hba-a2", "Hbb", "Alas2", "Gypa", "Ahsp"],
    "megakaryocyte": ["Ppbp", "Pf4", "Gp9", "Itga2b", "Nrgn"],
    "S_phase": ["Mcm2", "Mcm3", "Mcm4", "Mcm5", "Mcm6", "Mcm7", "Pcna", "Tyms", "Rrm1", "Rrm2", "Ung", "Pclaf"],
    "G2M_phase": ["Mki67", "Top2a", "Ube2c", "Cenpf", "Tpx2", "Nusap1", "Cdc20", "Ccnb1", "Ccnb2", "Birc5"],
    "interferon": ["Isg15", "Mx1", "Mx2", "Ifi6", "Ifi44", "Ifit1", "Ifit3", "Oas1a", "Usp18"],
    "activation_APC": ["Cd40", "Cd83", "Cd69", "Ccr7", "Rel", "Relb", "Nfkb1", "Nfkb2", "Ifi30"],
}

STATE_QC_RE = re.compile(
    r"^(mt-|rpl|rps|hist|h[1234]f|mki67|top2a|ube2c|cenp|tpx2|nusap1|pttg1|kif|"
    r"cdca|bub|sgo|hmmr|prc1|ndc80|dlgap5|cdc20|birc5|aspm|knstrn|tacc3|"
    r"ccna|ccnb|cdk1|plk1|aurk|pclaf|lig1|rrm|tyms|pcna|mcm|fos|jun|egr|hsp)",
    re.IGNORECASE,
)

ENSEMBL_GENE_ID_RE = re.compile(r"^ENS[A-Z0-9]*G\d+(?:\.\d+)?$", re.IGNORECASE)
LOC_GENE_ID_RE = re.compile(r"^LOC\d+$", re.IGNORECASE)
MITO_QC_RE = re.compile(
    r"^(?:MT-)?(?:COX[123]|ND[1-6]|ND4L|ATP6|ATP8|CYTB)$",
    re.IGNORECASE,
)


def naming_exclusion_reason(gene):
    """Return why a marker must not be used as a human-readable identity label."""
    symbol = str(gene).strip()
    if ENSEMBL_GENE_ID_RE.fullmatch(symbol):
        return "ensembl_stable_id_not_gene_symbol"
    if LOC_GENE_ID_RE.fullmatch(symbol):
        return "loc_placeholder_not_gene_symbol"
    if MITO_QC_RE.fullmatch(symbol):
        return "mitochondrial_qc_state_not_identity"
    if STATE_QC_RE.match(symbol):
        return "state_or_qc_gene_not_identity"
    return None


def cluster_sort_key(value):
    """Sort numeric cluster IDs by value, then nonnumeric IDs naturally."""
    text = str(value).strip()
    try:
        number = Decimal(text)
        if number.is_finite():
            return (0, number, text)
    except InvalidOperation:
        pass
    natural = tuple((0, int(part)) if part.isdigit() else (1, part.lower()) for part in re.split(r"(\d+)", text))
    return (1, natural, text)


def _shared_strings(zf):
    values = []
    with zf.open("xl/sharedStrings.xml") as handle:
        for _, elem in ET.iterparse(handle, events=("end",)):
            if elem.tag == NS + "si":
                values.append("".join(node.text or "" for node in elem.iter(NS + "t")))
                elem.clear()
    return values


def _column_index(cell_ref):
    match = re.match(r"[A-Z]+", cell_ref)
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    value = 0
    for char in match.group(0):
        value = value * 26 + ord(char) - 64
    return value - 1


def _parse_sheet_xml(path):
    rows, max_col = [], 0
    with zipfile.ZipFile(path) as zf:
        strings = _shared_strings(zf) if "xl/sharedStrings.xml" in zf.namelist() else []
        with zf.open("xl/worksheets/sheet1.xml") as handle:
            for _, elem in ET.iterparse(handle, events=("end",)):
                if elem.tag != NS + "row":
                    continue
                cells = {}
                for cell in elem.findall(NS + "c"):
                    idx = _column_index(cell.attrib["r"])
                    node = cell.find(NS + "v")
                    raw = None if node is None else node.text
                    if raw is not None and cell.attrib.get("t") == "s":
                        raw = strings[int(raw)]
                    cells[idx] = raw
                    max_col = max(max_col, idx + 1)
                rows.append(cells)
                elem.clear()
    return [[row.get(i) for i in range(max_col)] for row in rows]


def load_average(path):
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1 and len(rows[0]) > 1:
            return [list(row) for row in rows], "openpyxl"
    except Exception:
        pass
    return _parse_sheet_xml(path), "ooxml_fallback"


def _alias_index(headers, aliases):
    normalized = {str(value).strip().lower(): i for i, value in enumerate(headers)}
    for alias in aliases:
        if alias.lower() in normalized:
            return normalized[alias.lower()]
    raise ValueError(f"Missing required column. Expected one of: {aliases}")


def load_markers(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    headers = list(next(iterator))
    cluster_i = _alias_index(headers, ["Target_Cluster", "cluster", "seurat_clusters", "Cluster"])
    gene_i = _alias_index(headers, ["GeneName", "gene", "features", "Gene"])
    logfc_i = _alias_index(headers, ["log2FC", "avg_log2FC", "avg_logFC"])
    pct1_i = _alias_index(headers, ["pct.1", "pct1"])
    pct2_i = _alias_index(headers, ["pct.2", "pct2"])
    records = []
    for row in iterator:
        if not row[gene_i]:
            continue
        records.append({
            "cluster": str(row[cluster_i]), "gene": str(row[gene_i]),
            "log2FC": float(row[logfc_i] or 0), "pct1": float(row[pct1_i] or 0),
            "pct2": float(row[pct2_i] or 0),
        })
    return records


def classify_average_matrix(expression):
    """Classify whether average values look expression-like or gene-centered/scaled."""
    rows = [values for values in expression.values() if values]
    flattened = [float(value) for values in rows for value in values]
    if not flattened:
        return {
            "classification": "unknown_empty",
            "negative_fraction": 0.0,
            "median_abs_gene_mean": 0.0,
            "median_gene_sd": 0.0,
            "signature_score_policy": "disabled",
            "canonical_expression_policy": "do_not_use",
        }
    negative_fraction = sum(value < 0 for value in flattened) / len(flattened)
    gene_means = [statistics.fmean(values) for values in rows]
    gene_sds = [statistics.pstdev(values) for values in rows if len(values) > 1]
    median_abs_gene_mean = statistics.median(abs(value) for value in gene_means)
    median_gene_sd = statistics.median(gene_sds) if gene_sds else 0.0
    if negative_fraction >= 0.05 and median_abs_gene_mean <= 0.25:
        classification = "gene_centered_or_scaled"
        signature_policy = "mean_relative_scaled_value_prioritization_only"
        canonical_policy = "relative_rank_only_require_de_marker_support"
    elif negative_fraction >= 0.05:
        classification = "contains_negative_values_unknown_transform"
        signature_policy = "relative_priority_only"
        canonical_policy = "do_not_treat_as_absolute_expression"
    else:
        classification = "nonnegative_expression_like"
        signature_policy = "mean_log1p_nonnegative_expression"
        canonical_policy = "supporting_evidence_only"
    return {
        "classification": classification,
        "negative_fraction": round(negative_fraction, 4),
        "median_abs_gene_mean": round(median_abs_gene_mean, 4),
        "median_gene_sd": round(median_gene_sd, 4),
        "signature_score_policy": signature_policy,
        "canonical_expression_policy": canonical_policy,
    }


def _signature_score(expression, genes, cluster_index, matrix_semantics):
    lookup = {gene.upper(): vals for gene, vals in expression.items()}
    raw_values = [float(lookup[g.upper()][cluster_index]) for g in genes if g.upper() in lookup]
    if matrix_semantics["classification"] == "gene_centered_or_scaled":
        values = raw_values
    else:
        values = [math.log1p(max(value, 0.0)) for value in raw_values]
    return round(sum(values) / len(values), 4) if values else 0.0


def _signature_marker_support(ranked, genes, limit=100):
    wanted = {gene.upper() for gene in genes}
    hits = [record for record in ranked[:limit] if record["gene"].upper() in wanted and record["log2FC"] > 0]
    return {
        "count": len(hits),
        "genes": [record["gene"] for record in hits],
        "best_log2FC": round(max((record["log2FC"] for record in hits), default=0.0), 4),
    }


def build_evidence(avg_path, marker_path, top_n=60, informative_n=25, species=""):
    average, reader = load_average(Path(avg_path))
    if not average:
        raise ValueError("Average-expression file is empty")
    original_gene_header = str(average[0][0]).strip()
    accepted_gene_headers = {"genename", "gene", "features", "feature", "cluster"}
    if original_gene_header.lower() not in accepted_gene_headers:
        raise ValueError(
            "Average-expression column A must identify genes using one of: "
            "GeneName, Gene, features, feature, Cluster"
        )
    raw_clusters = [str(x) for x in average[0][1:]]
    cluster_order = sorted(range(len(raw_clusters)), key=lambda index: cluster_sort_key(raw_clusters[index]))
    clusters = [raw_clusters[index] for index in cluster_order]
    expression = {
        str(row[0]): [float(row[index + 1] or 0) for index in cluster_order]
        for row in average[1:] if row and row[0]
    }
    matrix_semantics = classify_average_matrix(expression)
    markers = load_markers(Path(marker_path))
    by_cluster = defaultdict(list)
    for record in markers:
        by_cluster[record["cluster"]].append(record)

    missing = [c for c in clusters if c not in by_cluster]
    extra = [c for c in by_cluster if c not in clusters]
    canonical = sorted({g.upper() for genes in SIGNATURES.values() for g in genes})
    expr_lookup = {g.upper(): (g, vals) for g, vals in expression.items()}
    canonical_expression = {expr_lookup[g][0]: expr_lookup[g][1] for g in canonical if g in expr_lookup}
    profiles = {}
    state_names = {"S_phase", "G2M_phase", "interferon", "activation_APC"}

    for cluster_index, cluster in enumerate(clusters):
        ranked = sorted(by_cluster.get(cluster, []), key=lambda x: (x["log2FC"], x["pct1"] - x["pct2"]), reverse=True)
        top = ranked[:top_n]
        naming_eligible = [r for r in ranked if naming_exclusion_reason(r["gene"]) is None]
        informative = naming_eligible[:informative_n]
        naming_top_marker = naming_eligible[0] if naming_eligible else None
        excluded_naming_markers = [
            {**record, "exclusion_reason": naming_exclusion_reason(record["gene"])}
            for record in top
            if naming_exclusion_reason(record["gene"]) is not None
        ]
        qc_fraction = sum(bool(STATE_QC_RE.match(r["gene"])) for r in top[:50]) / max(min(len(top), 50), 1)
        scores = {
            name: _signature_score(expression, genes, cluster_index, matrix_semantics)
            for name, genes in SIGNATURES.items()
        }
        marker_support = {
            name: _signature_marker_support(ranked, genes)
            for name, genes in SIGNATURES.items()
        }
        lineages = sorted(((k, v) for k, v in scores.items() if k not in state_names), key=lambda x: x[1], reverse=True)
        alerts = []
        if len(ranked) < 10:
            alerts.append("fewer_than_10_markers")
        if qc_fraction >= 0.6:
            alerts.append("state_or_qc_dominated")
        if naming_top_marker is None:
            alerts.append("no_standardized_informative_naming_marker")
        if len(lineages) >= 2 and lineages[0][1] > 0 and lineages[1][1] >= lineages[0][1] * 0.85:
            alerts.append("competing_lineage_signatures")
        if (
            matrix_semantics["classification"] != "nonnegative_expression_like"
            and lineages
            and marker_support[lineages[0][0]]["count"] < 2
        ):
            alerts.append("top_lineage_score_lacks_de_marker_support")
        profiles[cluster] = {
            "marker_count": len(ranked), "top_markers": top,
            "raw_top_marker": top[0] if top else None,
            "naming_top_marker": naming_top_marker,
            "excluded_naming_markers": excluded_naming_markers,
            "top_informative_markers": informative, "signature_scores": scores,
            "signature_marker_support": marker_support,
            "ranked_lineage_signatures": lineages[:5],
            "qc_state_fraction_top50": round(qc_fraction, 3), "alerts": alerts,
        }

    return {
        "schema_version": "2.4", "average_reader": reader,
        "average_gene_header": original_gene_header,
        "average_gene_header_normalized_to": "GeneName",
        "average_gene_names": list(expression),
        "average_shape": [len(average) - 1, len(clusters)], "clusters": clusters,
        "marker_cluster_ids": sorted(by_cluster, key=cluster_sort_key), "missing_marker_clusters": missing,
        "extra_marker_clusters": extra, "cluster_profiles": profiles,
        "canonical_expression_by_gene": canonical_expression,
        "average_matrix_semantics": matrix_semantics,
        "signature_definitions": SIGNATURES,
        "notes": [
            "Signature scores are prioritization aids, not final labels; interpret them according to average_matrix_semantics.",
            "Centered/scaled averages are relative ranks only and must not override positive DE-marker support.",
            "Consequential off-parent lineage calls require at least two coherent positive DE markers and negative-evidence review.",
            "Fallback naming uses naming_top_marker: the highest-log2FC standardized informative GeneName.",
            "Ensembl stable IDs such as ENSSSCG.../ENSG..., LOC placeholders, and state/QC-only genes are retained as evidence but excluded from identity labels.",
            "Cell-cycle/QC genes are state evidence and are removed from top_informative_markers and fallback identity naming.",
            "Plasma calls require a coherent PRDM1/XBP1/SDC1/MZB1/DERL3 program; JCHAIN alone is insufficient.",
            "Cluster IDs are normalized to numeric ascending order when numeric, otherwise natural alphanumeric order.",
        ],
        "naming_marker_policy": {
            "species": species,
            "rule": "highest_log2FC_standardized_informative_GeneName",
            "contextual_override": "first_defensible_top_informative_marker_with_complete_preceding_exclusion_audit",
            "excluded_patterns": [
                "Ensembl stable gene IDs: ENS...G<digits>",
                "LOC<digits> placeholders",
                "mitochondrial, ribosomal, histone, cell-cycle, immediate-early, and stress/QC-only genes",
            ],
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--avg", required=True)
    parser.add_argument("--markers", required=True)
    parser.add_argument("--output-json", required=True)
    parser.add_argument("--top-n", type=int, default=60)
    parser.add_argument("--informative-n", type=int, default=25)
    parser.add_argument("--species", default="")
    args = parser.parse_args()
    result = build_evidence(args.avg, args.markers, args.top_n, args.informative_n, args.species)
    output = Path(args.output_json)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "average_reader": result["average_reader"], "average_shape": result["average_shape"],
        "clusters": result["clusters"], "missing_marker_clusters": result["missing_marker_clusters"],
        "extra_marker_clusters": result["extra_marker_clusters"], "output_json": str(output.resolve()),
        "output_bytes": output.stat().st_size,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()




