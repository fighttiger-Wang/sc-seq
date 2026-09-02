#!/usr/bin/env python3
"""Regression test for the qualitative major-celltype workbook contract."""

import argparse
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


SKILL = Path(__file__).resolve()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--work-dir", required=True)
    args = parser.parse_args()
    work = Path(args.work_dir).resolve()
    work.mkdir(parents=True, exist_ok=True)
    repo = next(parent for parent in work.parents if (parent / "local-marketplace").is_dir()) / "local-marketplace"
    skill = repo / "plugins" / "sc-major-celltype-annotation-auto" / "skills" / "sc-major-celltype-annotation-auto"
    evidence = {
        "clusters": ["10", "2", "0"], "average_shape": [12, 3], "average_reader": "test",
        "confirmed_metadata": {"species": "Human", "tissue": "lung", "annotation_level": "major", "parent_population": "All_cells"},
        "source_paths": {"cell_avg_exp": "avg.tsv", "marker_table": "markers.xlsx"},
        "cluster_profiles": {
            "0": {"top_markers": [{"gene": "EPCAM", "mean_expr": 3.0, "pct1": 0.8, "pct2": 0.1, "log2FC": 2.0}]},
            "2": {"top_markers": [{"gene": "TRAC", "mean_expr": 2.5, "pct1": 0.7, "pct2": 0.05, "log2FC": 1.8}]},
            "10": {"top_markers": [{"gene": "NCR1", "mean_expr": 2.8, "pct1": 0.65, "pct2": 0.03, "log2FC": 2.1}]},
        },
        "qualitative_annotation_evidence": {
            "0": {"stable_id": "Epithelial_cell", "qualitative_gates": {"identity_anchor": "通过"}},
            "2": {"stable_id": "T_cell", "qualitative_gates": {"identity_anchor": "通过"}},
            "10": {"stable_id": "NK_cell", "qualitative_gates": {"identity_anchor": "通过"}},
        },
    }
    rows = [("10", "NK细胞", "NK_cell", "NCR1"), ("2", "T细胞", "T_cell", "TRAC"), ("0", "上皮细胞", "Epithelial_cell", "EPCAM")]
    records = [{
        "cluster_id": cluster, "celltype_cn": cn, "celltype_en": en, "stable_id": en,
        "broad_type": en, "supporting_markers": marker, "candidate_labels": en,
        "rationale": "Qualitative multi-marker program retained.",
        "review_action": "Retain identity and verify on UMAP.",
        "literature_source": "Curated Cell Ontology reference",
    } for cluster, cn, en, marker in rows]
    ep, rp, output = work / "evidence.json", work / "records.json", work / "major.xlsx"
    ep.write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    rp.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    completed = subprocess.run([
        sys.executable, str(skill / "scripts" / "build_annotation_workbook.py"),
        "--records", str(rp), "--evidence", str(ep), "--output", str(output),
        "--workspace-root", str(work.parents[1]), "--force",
    ], text=True, capture_output=True)
    if completed.returncode:
        raise RuntimeError(completed.stdout + completed.stderr)
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["绘图列表", "注释结果", "详细证据", "细胞类型与文献", "说明与数据来源"]
    assert [workbook["绘图列表"].cell(row, 1).value for row in range(2, 5)] == ["0", "2", "10"]
    headers = [cell.value for cell in workbook["注释结果"][1]]
    assert "质量评分" not in headers and "置信度" not in headers
    assert all(sheet.auto_filter.ref is None for sheet in workbook.worksheets)
    assert all(not cell.alignment.wrap_text for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row)
    qa = json.loads(output.with_suffix(".qa.json").read_text(encoding="utf-8"))
    assert qa["aggregate_scores_exported"] is False and qa["confidence_exported"] is False
    print(json.dumps({"status": "pass", "checks": 10, "output": str(output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
