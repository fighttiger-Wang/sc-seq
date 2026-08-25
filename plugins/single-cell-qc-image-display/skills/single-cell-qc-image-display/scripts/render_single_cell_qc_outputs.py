#!/usr/bin/env python3
"""Render single-cell QC rows to an Excel workbook and a presentation PNG."""

from __future__ import annotations

import argparse
import json
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


DEFAULT_COLUMNS = [
    "项目号",
    "客户名",
    "组学方案",
    "实验物种",
    "样本数",
    "测序量",
    "组织",
    "组织消化方案",
    "预期捕获细胞数",
    "关注细胞类型",
    "实验前细胞状态质控",
    "服务范畴",
    "Sample",
    "Estimated number",
    "Mean reads per cell",
    "Mean genes per cell",
    "Reads mapped to genome",
    "Sequencing saturation",
    "Cell type",
]

BASE_WIDTHS = {
    "项目号": 24,
    "客户名": 14,
    "组学方案": 36,
    "实验物种": 12,
    "样本数": 10,
    "测序量": 12,
    "组织": 12,
    "组织消化方案": 44,
    "预期捕获细胞数": 18,
    "关注细胞类型": 34,
    "实验前细胞状态质控": 22,
    "服务范畴": 14,
    "Sample": 24,
    "Estimated number": 18,
    "Mean reads per cell": 22,
    "Mean genes per cell": 22,
    "Reads mapped to genome": 24,
    "Sequencing saturation": 24,
    "Cell type": 62,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_json", help="JSON file containing columns and rows")
    parser.add_argument("--output-dir", default="outputs/single-cell-qc-table")
    parser.add_argument("--basename", default="single-cell-qc-projects")
    parser.add_argument("--title", help="Override image title")
    parser.add_argument("--subtitle", help="Override image subtitle")
    return parser.parse_args()


def load_payload(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("input JSON must be an object")
    rows = payload.get("rows")
    if not isinstance(rows, list) or not rows:
        raise ValueError("input JSON must include a non-empty rows array")
    return payload


def normalize_columns(payload: dict[str, Any]) -> list[str]:
    columns = payload.get("columns")
    if isinstance(columns, list) and columns:
        return [str(c) for c in columns]
    present = OrderedDict((c, None) for c in DEFAULT_COLUMNS)
    for row in payload["rows"]:
        if isinstance(row, dict):
            for key in row:
                present.setdefault(str(key), None)
    return list(present.keys())


def rows_as_matrix(rows: list[dict[str, Any]], columns: list[str]) -> list[list[Any]]:
    matrix = []
    for row in rows:
        matrix.append([row.get(col, "") if isinstance(row, dict) else "" for col in columns])
    return matrix


def make_excel(payload: dict[str, Any], columns: list[str], out_path: Path) -> None:
    rows = payload["rows"]
    matrix = rows_as_matrix(rows, columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "单细胞QC汇总"
    ws.append(columns)
    for row in matrix:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=10)
    body_font = Font(name="Microsoft YaHei", color="1F2933", size=10)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 34

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row[0].row].height = 54

    for idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = BASE_WIDTHS.get(col_name, 18)

    ws.freeze_panes = "A2"
    ref = f"A1:{get_column_letter(len(columns))}{len(matrix) + 1}"
    table = Table(displayName="SingleCellQCTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    add_project_overview(wb, rows)
    wb.save(out_path)


def add_project_overview(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    project_fields = ["项目号", "客户名", "样本数", "组织", "组织消化方案", "关注细胞类型"]
    projects: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row in rows:
        pid = str(row.get("项目号", "")).strip()
        if not pid:
            continue
        entry = projects.setdefault(pid, {field: row.get(field, "") for field in project_fields})
        sample = row.get("Sample", "")
        if sample:
            entry.setdefault("样本列表", []).append(str(sample))
    if not projects:
        return

    ws = wb.create_sheet("项目概览")
    headers = project_fields + ["样本列表"]
    ws.append(headers)
    for entry in projects.values():
        ws.append([entry.get(field, "") for field in project_fields] + [", ".join(entry.get("样本列表", []))])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", color="1F2933", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row[0].row].height = 46
    widths = [24, 14, 10, 12, 44, 34, 42]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ref = f"A1:{get_column_letter(len(headers))}{len(projects) + 1}"
    table = Table(displayName="ProjectOverviewTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/simsun.ttc",
        "C:/Windows/Fonts/arial.ttf",
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size)
    return ImageFont.load_default()


def text_size(draw: ImageDraw.ImageDraw, text: str, fnt: ImageFont.ImageFont) -> tuple[int, int]:
    box = draw.textbbox((0, 0), str(text), font=fnt)
    return box[2] - box[0], box[3] - box[1]


def wrap_text(draw: ImageDraw.ImageDraw, text: Any, fnt: ImageFont.ImageFont, max_width: int) -> list[str]:
    text = "" if text is None else str(text)
    if not text:
        return [""]
    lines: list[str] = []
    current = ""
    for ch in text:
        candidate = current + ch
        if text_size(draw, candidate, fnt)[0] <= max_width or not current:
            current = candidate
        else:
            lines.append(current)
            current = ch
    if current:
        lines.append(current)
    return lines or [""]


def draw_wrapped(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: Any,
    fnt: ImageFont.ImageFont,
    max_width: int,
    max_height: int,
    fill: str = "#102A43",
    align: str = "left",
    line_gap: int = 6,
) -> None:
    x, y = xy
    lines = wrap_text(draw, text, fnt, max_width)
    line_h = text_size(draw, "Hg", fnt)[1] + line_gap
    max_lines = max(1, max_height // line_h)
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        lines[-1] = lines[-1].rstrip(" .,:;，。、") + "..."
    total_h = len(lines) * line_h - line_gap
    cy = y + max(0, (max_height - total_h) // 2)
    for line in lines:
        tw, _ = text_size(draw, line, fnt)
        if align == "center":
            tx = x + (max_width - tw) / 2
        elif align == "right":
            tx = x + max_width - tw
        else:
            tx = x
        draw.text((tx, cy), line, font=fnt, fill=fill)
        cy += line_h


def image_column_widths(columns: list[str]) -> list[int]:
    default_px = {
        "项目号": 300,
        "客户名": 130,
        "组学方案": 420,
        "实验物种": 110,
        "样本数": 90,
        "测序量": 110,
        "组织": 100,
        "组织消化方案": 520,
        "预期捕获细胞数": 180,
        "关注细胞类型": 360,
        "实验前细胞状态质控": 220,
        "服务范畴": 120,
        "Sample": 220,
        "Estimated number": 190,
        "Mean reads per cell": 220,
        "Mean genes per cell": 210,
        "Reads mapped to genome": 235,
        "Sequencing saturation": 245,
        "Cell type": 650,
    }
    return [default_px.get(col, 220) for col in columns]


def make_png(payload: dict[str, Any], columns: list[str], out_path: Path, title: str, subtitle: str) -> None:
    rows = payload["rows"]
    matrix = rows_as_matrix(rows, columns)
    widths = image_column_widths(columns)

    margin = 70
    table_width = sum(widths)
    title_h = 150
    card_h = 190
    gap = 36
    header_h = 96
    row_h = 132
    footer_h = 76
    width = table_width + margin * 2
    height = margin + title_h + card_h + gap + header_h + row_h * len(matrix) + footer_h

    img = Image.new("RGB", (width, height), "#F4F7FB")
    draw = ImageDraw.Draw(img)
    title_font = font(58, True)
    subtitle_font = font(26)
    card_label_font = font(24, True)
    card_value_font = font(42, True)
    header_font = font(22, True)
    cell_font = font(21)
    small_font = font(18)

    draw.text((margin, margin), title, font=title_font, fill="#173B57")
    draw.text((margin, margin + 76), subtitle, font=subtitle_font, fill="#52606D")

    project_ids = [str(r.get("项目号", "")).strip() for r in rows if str(r.get("项目号", "")).strip()]
    projects = list(OrderedDict((pid, None) for pid in project_ids).keys())
    species = sorted({str(r.get("实验物种", "")).strip() for r in rows if str(r.get("实验物种", "")).strip()})
    seq = sorted({str(r.get("测序量", "")).strip() for r in rows if str(r.get("测序量", "")).strip()})
    digestion_count = sum(1 for pid in projects if any(str(r.get("项目号", "")) == pid and str(r.get("组织消化方案", "")).strip() for r in rows))
    cards = [
        ("项目数", str(len(projects)), "LC-X 项目"),
        ("样本数", str(len(rows)), "Sample 明细"),
        ("已补充消化方案", str(digestion_count), "按项目号填充"),
        ("物种", "、".join(species) or "", "项目字段"),
        ("测序量", "、".join(seq) or "", "项目字段"),
    ]
    card_gap = 24
    card_width = (table_width - card_gap * (len(cards) - 1)) // len(cards)
    card_y = margin + title_h
    for i, (label, value, note) in enumerate(cards):
        x = margin + i * (card_width + card_gap)
        draw.rounded_rectangle((x, card_y, x + card_width, card_y + card_h), radius=24, fill="#FFFFFF", outline="#E2E8F0", width=2)
        draw.text((x + 28, card_y + 26), label, font=card_label_font, fill="#52606D")
        draw.text((x + 28, card_y + 72), value, font=card_value_font, fill="#1F4E78")
        draw.text((x + 28, card_y + 132), note, font=subtitle_font, fill="#627D98")

    table_x = margin
    table_y = margin + title_h + card_h + gap
    draw.rounded_rectangle((table_x - 8, table_y - 8, table_x + table_width + 8, table_y + header_h + row_h * len(matrix) + 8), radius=18, fill="#E8EEF6")

    x = table_x
    for header, col_width in zip(columns, widths):
        draw.rectangle((x, table_y, x + col_width, table_y + header_h), fill="#1F4E78", outline="#CBD5E1")
        draw_wrapped(draw, (x + 8, table_y + 8), header, header_font, col_width - 16, header_h - 16, fill="#FFFFFF", align="center")
        x += col_width

    palette = ["#EAF4FF", "#EEF8F0", "#FFF4E5", "#F4F0FF", "#FEEEEE", "#EEF2FF"]
    project_fill = {pid: palette[i % len(palette)] for i, pid in enumerate(projects)}
    center_names = {"客户名", "实验物种", "样本数", "测序量", "组织", "预期捕获细胞数", "实验前细胞状态质控", "服务范畴", "Estimated number", "Mean reads per cell", "Mean genes per cell", "Reads mapped to genome", "Sequencing saturation"}
    small_names = {"组学方案", "组织消化方案", "关注细胞类型", "Cell type"}

    for row_idx, values in enumerate(matrix):
        y = table_y + header_h + row_idx * row_h
        pid = str(values[columns.index("项目号")]) if "项目号" in columns else ""
        base_fill = "#FFFFFF" if row_idx % 2 == 0 else "#F8FBFF"
        x = table_x
        for col_name, value, col_width in zip(columns, values, widths):
            fill = project_fill.get(pid, base_fill) if col_name == "项目号" else base_fill
            draw.rectangle((x, y, x + col_width, y + row_h), fill=fill, outline="#CBD5E1")
            align = "center" if col_name in center_names else "left"
            fnt = small_font if col_name in small_names else cell_font
            draw_wrapped(draw, (x + 8, y + 8), value, fnt, col_width - 16, row_h - 16, align=align)
            x += col_width

    missing = []
    if "项目号" in columns and "组织消化方案" in columns:
        for pid in projects:
            if not any(str(r.get("项目号", "")) == pid and str(r.get("组织消化方案", "")).strip() for r in rows):
                missing.append(pid)
    note = "注：一行代表一个 Sample；颜色区分不同项目号。"
    if missing:
        note += " 未捕获组织消化方案：" + "、".join(missing) + "。"
    draw.text((margin, table_y + header_h + row_h * len(matrix) + 28), note, font=subtitle_font, fill="#52606D")

    img.save(out_path, "PNG", optimize=True)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_json)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    payload = load_payload(input_path)
    columns = normalize_columns(payload)
    title = args.title or payload.get("title") or "单细胞转录组 QC 项目汇总"
    subtitle = args.subtitle or payload.get("subtitle") or "按项目号集中展示项目基础信息、样本 QC 指标与细胞类型注释"
    xlsx_path = output_dir / f"{args.basename}.xlsx"
    png_path = output_dir / f"{args.basename}-big-image.png"
    make_excel(payload, columns, xlsx_path)
    make_png(payload, columns, png_path, str(title), str(subtitle))
    print(json.dumps({"xlsx": str(xlsx_path.resolve()), "png": str(png_path.resolve()), "rows": len(payload["rows"])}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
